"""표준 구매요청서(139) + 표준 견적서(37) → RAG 청크 시드
기존 non-faq 청크를 삭제하고, Excel 내용을 마크다운 청크로 변환하여
knowledge_chunks 테이블에 임베딩과 함께 삽입한다.

사용법:
  cd backend
  python -m scripts.seed_rag_v3
  python -m scripts.seed_rag_v3 --dry-run   (DB 저장 없이 미리보기)
"""
import sys, os, io, time, re, glob, argparse
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from dotenv import load_dotenv
load_dotenv()

import openpyxl
from app.db.supabase_client import get_client
from app.rag.embedder import embed_document

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
PR_EXCEL_DIR = os.path.join(PROJECT_ROOT, "표준 구매요청서")
RFQ_EXCEL_DIR = os.path.join(PROJECT_ROOT, "표준 견적서")

# L1 폴더 → 대분류명
L1_MAP = {
    "L01": "사무·총무", "L02": "인사·복리후생", "L03": "시설·건물관리",
    "L04": "차량·출장", "L05": "보험 서비스", "L06": "전문용역·컨설팅",
    "L07": "마케팅", "L08": "IT/ICT", "L09": "물류",
    "L10": "생산관리", "L11": "연구개발",
}

EXCLUDE_COLS = ["시장렌탈료참고", "B2B예산단가참고", "시장렌탈참고", "B2B예산"]


def get_l1_from_folder(folder_name: str) -> str:
    m = re.match(r'^(L\d{2})', folder_name)
    return L1_MAP.get(m.group(1), folder_name) if m else folder_name


def extract_prefix(filename: str) -> str:
    m = re.match(r'^([LMR]\d{4})', filename)
    return m.group(1) if m else ""


def extract_service_name(filename: str, prefix: str) -> str:
    name = filename.replace(prefix + "_", "")
    name = re.sub(r'_(구매요청서|표준견적서[^.]*)?\.xlsx$', '', name)
    return name.replace("_", " ").strip()


# ─────────────────────────────────────
# 구매요청서 청킹 (Sheet 1~3)
# ─────────────────────────────────────
def chunk_pr_excel(filepath: str, l1_name: str) -> list[dict]:
    """구매요청서 Excel → 청크 리스트"""
    filename = os.path.basename(filepath)
    prefix = extract_prefix(filename)
    svc_name = extract_service_name(filename, prefix)

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception:
        return []

    chunks = []

    # ── Sheet 1: 구매요청서 (메인 청크) ──
    ws = wb[wb.sheetnames[0]]
    lines = [f"# {svc_name} — 구매요청서", f"분류: {l1_name} > {svc_name}", ""]

    required_items = []
    optional_items = []

    for r in range(9, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        if c1 is None:
            continue
        c1_str = str(c1).strip()

        if c1_str.startswith("▶"):
            section = c1_str.replace("▶", "").strip()
            required_items.append(f"\n### {section}")
            continue

        try:
            int(c1_str)
        except ValueError:
            continue

        name = str(ws.cell(r, 3).value or "").strip()
        desc = str(ws.cell(r, 4).value or "").replace("\n", " ").strip()[:120]
        req = str(ws.cell(r, 5).value or "").strip()
        impact = str(ws.cell(r, 6).value or "").strip()

        if not name:
            continue

        entry = f"- **{name}**: {desc}" if desc else f"- **{name}**"
        if "상" in impact:
            entry += " ⚡단가영향 상"

        if "필수" in req:
            required_items.append(entry)
        else:
            optional_items.append(entry)

    if required_items:
        lines.append("## 필수 항목")
        lines.extend(required_items)
    if optional_items:
        lines.append("\n## 옵션 항목")
        lines.extend(optional_items)

    chunk_text = "\n".join(lines)
    if len(chunk_text) > 50:
        chunks.append({
            "content": chunk_text,
            "category": l1_name,
            "doc_name": f"{l1_name}/{svc_name}/구매요청서",
            "chunk_index": 0,
            "metadata": {"l1": l1_name, "l3_code": prefix, "source": "표준구매요청서", "version": "0326"},
        })

    # ── Sheet 2: 단가영향 요소 (보조 청크) ──
    if len(wb.sheetnames) >= 2:
        ws2 = wb[wb.sheetnames[1]]
        lines2 = [f"# {svc_name} — 단가영향 요소", f"분류: {l1_name}", ""]
        for r in range(1, min(ws2.max_row + 1, 30)):
            row_vals = []
            for c in range(1, min(ws2.max_column + 1, 8)):
                v = ws2.cell(r, c).value
                if v:
                    row_vals.append(str(v).replace("\n", " ").strip()[:60])
            if row_vals:
                lines2.append(" | ".join(row_vals))

        text2 = "\n".join(lines2)
        if len(text2) > 50:
            chunks.append({
                "content": text2[:2000],
                "category": l1_name,
                "doc_name": f"{l1_name}/{svc_name}/단가영향",
                "chunk_index": 1,
                "metadata": {"l1": l1_name, "l3_code": prefix, "source": "표준구매요청서", "version": "0326"},
            })

    wb.close()
    return chunks


# ─────────────────────────────────────
# 견적서 청킹 (Sheet 1~5)
# ─────────────────────────────────────
SHEET_LABELS = ["가격견적", "계약조건", "부가서비스", "SLA품질", "TCO정산"]


def chunk_rfq_excel(filepath: str, l1_name: str) -> list[dict]:
    """견적서 Excel → 청크 리스트 (소싱담당자 Zone 1-2만)"""
    filename = os.path.basename(filepath)
    prefix = extract_prefix(filename)
    svc_name = extract_service_name(filename, prefix)

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception:
        return []

    chunks = []

    for si in range(min(5, len(wb.sheetnames))):
        ws = wb[wb.sheetnames[si]]
        label = SHEET_LABELS[si] if si < len(SHEET_LABELS) else f"시트{si+1}"

        lines = [f"# {svc_name} — 견적서 {label}", f"분류: {l1_name}", ""]

        # 헤더 행 찾기
        header_row = 9
        for r in range(7, min(ws.max_row + 1, 12)):
            if ws.cell(r, 1).value and str(ws.cell(r, 1).value).strip() == "No.":
                header_row = r
                break

        # 소싱담당자 영역 컬럼 감지 (공급사 영역 시작 위치)
        supplier_start = ws.max_column + 1
        for c in range(1, ws.max_column + 1):
            for check_row in [3, 8]:
                v = ws.cell(check_row, c).value
                if v and "공급사" in str(v) and "필수" in str(v):
                    supplier_start = min(supplier_start, c)

        # 데이터 파싱
        for r in range(header_row + 1, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            if c1 is None:
                continue
            c1_str = str(c1).strip()

            if c1_str.startswith("▶"):
                lines.append(f"\n### {c1_str.replace('▶', '').strip()}")
                continue
            if c1_str.startswith("🔶") or c1_str.startswith("발주사"):
                continue

            try:
                int(c1_str)
            except ValueError:
                if not re.match(r'^[①②③④⑤⑥⑦⑧⑨⑩]', c1_str):
                    continue

            # 소싱담당자 영역만 (공급사 시작 전까지)
            vals = []
            for c in range(2, min(supplier_start, ws.max_column + 1)):
                hdr = ws.cell(header_row, c).value
                if hdr and any(kw in str(hdr) for kw in EXCLUDE_COLS):
                    continue
                v = ws.cell(r, c).value
                if v:
                    vals.append(str(v).replace("\n", " ").strip()[:80])

            if vals:
                lines.append("- " + " | ".join(vals[:5]))

        text = "\n".join(lines)
        if len(text) > 50:
            chunks.append({
                "content": text[:2000],
                "category": l1_name,
                "doc_name": f"{l1_name}/{svc_name}/견적서-{label}",
                "chunk_index": si,
                "metadata": {"l1": l1_name, "l3_code": prefix, "source": "표준견적서", "version": "0326"},
            })

    wb.close()
    return chunks


# ─────────────────────────────────────
# 기존 청크 삭제
# ─────────────────────────────────────
def delete_existing_chunks(supabase) -> int:
    existing = supabase.table("knowledge_chunks").select("id").neq("category", "faq").execute()
    if not existing.data:
        return 0
    ids = [r["id"] for r in existing.data]
    for i in range(0, len(ids), 50):
        supabase.table("knowledge_chunks").delete().in_("id", ids[i:i+50]).execute()
    return len(ids)


# ─────────────────────────────────────
# 메인
# ─────────────────────────────────────
def run():
    if sys.platform == "win32":
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace", line_buffering=True)

    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    all_chunks = []

    # 1. 구매요청서 139개
    if os.path.isdir(PR_EXCEL_DIR):
        print(f"=== 표준 구매요청서 청킹 ===")
        for folder in sorted(os.listdir(PR_EXCEL_DIR)):
            fp = os.path.join(PR_EXCEL_DIR, folder)
            if not os.path.isdir(fp):
                continue
            l1 = get_l1_from_folder(folder)
            for fpath in sorted(glob.glob(os.path.join(fp, "*.xlsx"))):
                chunks = chunk_pr_excel(fpath, l1)
                all_chunks.extend(chunks)
        print(f"  구매요청서 청크: {len(all_chunks)}개")

    # 2. 견적서 37개
    rfq_start = len(all_chunks)
    if os.path.isdir(RFQ_EXCEL_DIR):
        print(f"=== 표준 견적서 청킹 ===")
        for folder in sorted(os.listdir(RFQ_EXCEL_DIR)):
            fp = os.path.join(RFQ_EXCEL_DIR, folder)
            if not os.path.isdir(fp):
                continue
            l1 = get_l1_from_folder(folder)
            for fpath in sorted(glob.glob(os.path.join(fp, "*.xlsx"))):
                chunks = chunk_rfq_excel(fpath, l1)
                all_chunks.extend(chunks)
        print(f"  견적서 청크: {len(all_chunks) - rfq_start}개")

    print(f"\n총 청크: {len(all_chunks)}개")

    if args.dry_run:
        print("\n[DRY RUN] 처음 3개 청크:")
        for i, c in enumerate(all_chunks[:3]):
            print(f"\n--- 청크 {i+1}: {c['doc_name']} ---")
            print(c["content"][:300])
        print(f"\n[DRY RUN] 총 {len(all_chunks)}개 생성 예정")
        return

    # 3. DB 시드
    supabase = get_client()
    deleted = delete_existing_chunks(supabase)
    print(f"기존 청크 {deleted}개 삭제 완료")

    inserted = 0
    errors = 0
    for idx, chunk in enumerate(all_chunks):
        try:
            embedding = embed_document(chunk["content"])
            supabase.table("knowledge_chunks").insert({
                "content": chunk["content"],
                "embedding": embedding,
                "category": chunk["category"],
                "doc_name": chunk["doc_name"],
                "chunk_index": chunk["chunk_index"],
                "metadata": chunk["metadata"],
            }).execute()
            inserted += 1

            if (idx + 1) % 20 == 0:
                print(f"  {idx+1}/{len(all_chunks)} ({inserted} 삽입, {errors} 오류)")
            time.sleep(0.15)
        except Exception as e:
            errors += 1
            print(f"  [오류] {chunk['doc_name']}: {e}")
            time.sleep(1)

    print(f"\n[DONE] RAG v3 시드 완료: {inserted}개 삽입, {errors}개 오류")


if __name__ == "__main__":
    run()
