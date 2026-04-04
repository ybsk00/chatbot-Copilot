"""계약서 템플릿 시드 - 표준 계약서 DOCX 12개 + L3 매핑표 -> contract_templates 테이블

실행: python -m scripts.seed_contract_templates
"""
import sys, os, re, json, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from dotenv import load_dotenv
load_dotenv()

from docx import Document
import openpyxl
from app.db.supabase_client import get_client

# ── 경로 ──
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
CONTRACT_DIR = os.path.join(PROJECT_ROOT, "04 표준 계약서")
MAPPING_XLSX = os.path.join(CONTRACT_DIR, "L3_계약유형_매핑표.xlsx")

# ── 계약유형 코드→DOCX 파일명 매핑 ──
TYPE_FILES = {
    "A": "Type_A_물품구매계약서.docx",
    "B": "Type_B_일반용역계약서.docx",
    "C": "Type_C_시설공사계약서.docx",
    "D": "Type_D_렌탈리스계약서.docx",
    "E": "Type_E_SW라이선스클라우드계약서.docx",
    "F": "Type_F_IT개발SI용역계약서.docx",
    "G": "Type_G_전문컨설팅계약서.docx",
    "H": "Type_H_광고마케팅대행계약서.docx",
    "I": "Type_I_물류운송계약서.docx",
    "J": "Type_J_보험계약서.docx",
    "K": "Type_K_RnD개발용역계약서.docx",
    "L": "Type_L_지식재산로열티계약서.docx",
}

# ── PR 필드 자동매핑 힌트 (핵심 조항 입력필드 → PR c필드) ──
# 계약유형별로 조항의 입력필드가 PR 어디에 대응하는지 정의
PR_MAP_HINTS = {
    "품명": "c6", "규격": "c6", "서비스명": "c6", "품목명": "c6",
    "수량": "c10", "단위": "c10", "규모": "c10", "대수": "c10",
    "계약기간": "c9", "기간": "c9", "시작일": "c9", "종료일": "c9",
    "개발기간": "c9", "수행기간": "c9", "공사기간": "c9",
    "계약금액": None, "금액": None, "도급금액": None,  # 사용자 직접 입력
    "부가세": None, "부가가치세": None,
    "납품기한": None, "납품장소": None,
    "보증기간": None, "보증금": None, "하자보증": None,
    "선급금": None, "위약금": None,
    "지급조건": "c17", "지급방식": "c17", "결제": "c17",
    "렌탈료": None, "보험료": None, "로열티": None,
    "특약": None,
}


def guess_pr_map(field_text: str) -> str | None:
    """입력 필드 텍스트에서 PR 매핑 힌트 추출."""
    for keyword, pr_key in PR_MAP_HINTS.items():
        if keyword in field_text:
            return pr_key
    return None


def parse_contract_docx(filepath: str) -> dict:
    """계약서 DOCX 파싱 → 전체 조항 + 핵심 조항(입력필드 있는 것) 추출."""
    doc = Document(filepath)

    all_articles = []      # 전체 조항
    current_article = None
    current_body = []

    for p in doc.paragraphs:
        text = p.text.strip()
        if not text:
            continue

        # 조항 감지: 제N조 [제목]
        m = re.match(r'(제\d+조)\s*\[(.+?)\](.*)', text)
        if m:
            # 이전 조항 저장
            if current_article:
                current_article["body"] = "\n".join(current_body)
                all_articles.append(current_article)

            star = "★" in text
            current_article = {
                "num": m.group(1),
                "title": m.group(2),
                "star": star,
                "input_fields": [],
                "has_blank": False,
                "body_lines": [],
            }
            current_body = []

            # 제목 줄에 입력필드/빈칸 확인
            rest = m.group(3)
            fields = re.findall(r'【(.+?)】', rest)
            current_article["input_fields"].extend(fields)
            if '________' in rest or '____' in rest:
                current_article["has_blank"] = True
            continue

        if current_article:
            # 입력 필드 감지
            fields = re.findall(r'【(.+?)】', text)
            if fields:
                current_article["input_fields"].extend(fields)
            if '________' in text or '____' in text:
                current_article["has_blank"] = True

            # 조항 본문 수집 (▶ 해설은 제외)
            if not text.startswith("▶"):
                current_body.append(text)

    # 마지막 조항 저장
    if current_article:
        current_article["body"] = "\n".join(current_body)
        all_articles.append(current_article)

    # ── 핵심 조항 = 입력필드 또는 빈칸이 있는 조항 ──
    key_articles = []
    field_counter = 1

    for art in all_articles:
        is_key = bool(art["input_fields"]) or art["has_blank"]
        if not is_key:
            continue

        # 핵심 조항에서 입력 필드 구조화
        structured_fields = {}
        for raw_field in art["input_fields"]:
            field_key = f"k{field_counter}"
            # 입력 필드 텍스트 분석
            pr_map = guess_pr_map(raw_field)

            field_def = {
                "label": raw_field[:80],
                "type": "text",
                "required": True,
                "pr_map": pr_map,
            }

            # select 타입 감지
            if "포함/별도" in raw_field:
                field_def["type"] = "select"
                field_def["options"] = ["포함", "별도"]
            elif "일시불/분할" in raw_field:
                field_def["type"] = "select"
                field_def["options"] = ["일시불", "분할"]

            structured_fields[field_key] = field_def
            field_counter += 1

        # 빈칸만 있고 【】 필드가 없는 경우 → 본문에서 빈칸 추출
        if art["has_blank"] and not art["input_fields"]:
            # 본문에서 ________ 패턴의 줄 추출
            for line in art.get("body", "").split("\n"):
                if "________" in line or "____" in line:
                    field_key = f"k{field_counter}"
                    label = re.sub(r'[①②③④⑤⑥⑦⑧⑨⑩]\s*', '', line)
                    label = label.replace("________", "___").strip()[:80]
                    structured_fields[field_key] = {
                        "label": label,
                        "type": "text",
                        "required": True,
                        "pr_map": guess_pr_map(label),
                    }
                    field_counter += 1

        # 조항 본문 요약 (첫 2항만)
        body_text = art.get("body", "")
        body_lines = [l for l in body_text.split("\n") if l.strip()]
        summary_lines = body_lines[:3]
        summary = "\n".join(summary_lines)
        if len(summary) > 300:
            summary = summary[:300] + "..."

        key_articles.append({
            "num": art["num"],
            "title": art["title"],
            "star": art["star"],
            "summary": summary,
            "fields": structured_fields,
        })

    # ── 전체 조항 (미리보기/PDF용) ──
    formatted_articles = []
    for art in all_articles:
        formatted_articles.append({
            "num": art["num"],
            "title": art["title"],
            "star": art["star"],
            "body": art.get("body", ""),
            "is_key": bool(art["input_fields"]) or art["has_blank"],
        })

    return {
        "total_articles": len(all_articles),
        "key_articles": key_articles,
        "all_articles": formatted_articles,
    }


def load_l3_mapping() -> dict[str, list[str]]:
    """L3_계약유형_매핑표.xlsx → {계약유형코드: [L3코드 리스트]}"""
    wb = openpyxl.load_workbook(MAPPING_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]

    mapping = {}  # "A" → ["L3-010101", "L3-010102", ...]
    for row in range(2, ws.max_row + 1):
        l3_code = str(ws.cell(row, 4).value or "").strip()    # D열: L3 코드
        ct_code = str(ws.cell(row, 6).value or "").strip()    # F열: 계약유형코드
        if l3_code and ct_code:
            mapping.setdefault(ct_code, []).append(l3_code)

    return mapping


# ── 당사자 공통 필드 정의 ──
COMMON_FIELDS = {
    "buyer_name":  {"label": "발주자 상호", "type": "text", "required": True, "pr_map": "c1", "group": "buyer"},
    "buyer_rep":   {"label": "발주자 대표자", "type": "text", "required": True, "pr_map": None, "group": "buyer"},
    "buyer_addr":  {"label": "발주자 주소", "type": "text", "required": False, "pr_map": None, "group": "buyer"},
    "buyer_brn":   {"label": "발주자 사업자등록번호", "type": "text", "required": False, "pr_map": None, "group": "buyer"},
    "buyer_contact": {"label": "발주자 담당자", "type": "text", "required": False, "pr_map": "c3", "group": "buyer"},
    "buyer_phone": {"label": "발주자 연락처", "type": "text", "required": False, "pr_map": "c4", "group": "buyer"},
    "buyer_email": {"label": "발주자 이메일", "type": "text", "required": False, "pr_map": "c5", "group": "buyer"},
    "supplier_name": {"label": "수주자 상호", "type": "text", "required": True, "supplier_map": "company", "group": "supplier"},
    "supplier_rep":  {"label": "수주자 대표자", "type": "text", "required": True, "pr_map": None, "group": "supplier"},
    "supplier_addr": {"label": "수주자 주소", "type": "text", "required": False, "pr_map": None, "group": "supplier"},
    "supplier_brn":  {"label": "수주자 사업자등록번호", "type": "text", "required": False, "pr_map": None, "group": "supplier"},
}


def main():
    sb = get_client()

    # 1) L3 매핑 로드
    print("=== L3→계약유형 매핑표 로드 ===")
    l3_mapping = load_l3_mapping()
    total_l3 = sum(len(v) for v in l3_mapping.values())
    print(f"  {len(l3_mapping)}개 유형, {total_l3}개 L3 매핑")
    for ct, codes in sorted(l3_mapping.items()):
        print(f"  {ct}: {len(codes)}개 L3")

    # 2) DOCX 12개 파싱
    print("\n=== DOCX 파싱 ===")
    templates = []
    for ct_code, filename in sorted(TYPE_FILES.items()):
        filepath = os.path.join(CONTRACT_DIR, filename)
        if not os.path.exists(filepath):
            print(f"  [SKIP] 파일 없음: {filename}")
            continue

        print(f"  {ct_code} | {filename}")
        parsed = parse_contract_docx(filepath)
        contract_name = filename.replace(".docx", "").split("_", 2)[-1]

        print(f"    전체 조항: {parsed['total_articles']}개")
        print(f"    핵심 조항: {len(parsed['key_articles'])}개")
        for ka in parsed['key_articles']:
            star = " ★" if ka['star'] else ""
            fcount = len(ka['fields'])
            print(f"      {ka['num']} [{ka['title']}]{star} — 입력필드 {fcount}개")

        l3_codes = l3_mapping.get(ct_code, [])

        templates.append({
            "contract_type": ct_code,
            "contract_name": contract_name,
            "key_articles": parsed["key_articles"],
            "all_articles": parsed["all_articles"],
            "common_fields": COMMON_FIELDS,
            "l3_codes": l3_codes,
            "total_articles": parsed["total_articles"],
        })

    # 3) DB 시딩
    print(f"\n=== DB 시딩 ({len(templates)}개) ===")

    # 기존 데이터 삭제
    try:
        sb.table("contract_templates").delete().neq("id", 0).execute()
        print("  기존 contract_templates 삭제 완료")
    except Exception as e:
        print(f"  기존 삭제 실패 (테이블 없을 수 있음): {e}")

    for tpl in templates:
        row = {
            "contract_type": tpl["contract_type"],
            "contract_name": tpl["contract_name"],
            "key_articles": tpl["key_articles"],
            "all_articles": tpl["all_articles"],
            "common_fields": tpl["common_fields"],
            "l3_codes": tpl["l3_codes"],
            "total_articles": tpl["total_articles"],
        }
        try:
            sb.table("contract_templates").insert(row).execute()
            print(f"  ✓ {tpl['contract_type']} {tpl['contract_name']} — 핵심{len(tpl['key_articles'])}조 / L3 {len(tpl['l3_codes'])}개")
        except Exception as e:
            print(f"  ✗ {tpl['contract_type']} 시딩 실패: {e}")

    print("\n=== 완료 ===")
    print(f"  총 {len(templates)}개 계약서 템플릿 시딩")
    print(f"  총 {total_l3}개 L3 매핑")


if __name__ == "__main__":
    main()
