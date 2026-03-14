"""신규DB Excel → pr_schemas.py 자동 생성 스크립트"""
import sys
import io
import os
import re

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl

EXCEL_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "..", "신규db")

FILE_MAP = {
    "공기청정기_렌탈서비스_구매요청서.xlsx": ("건물 관리", "공기청정기 렌탈 서비스", "air_purifier_rental"),
    "디지털광고제작운영_구매요청서.xlsx": ("마케팅", "디지털 광고 제작 및 운영", "digital_ad"),
    "문서파기_서비스_구매요청서.xlsx": ("건물 관리", "문서파기 서비스", "document_shredding"),
    "물리보안_서비스_구매요청서.xlsx": ("건물 관리", "물리보안 서비스", "physical_security"),
    "바이럴마케팅대행_구매요청서.xlsx": ("마케팅", "바이럴 마케팅 대행", "viral_marketing"),
    "방역소독_구독서비스_구매요청서.xlsx": ("건물 관리", "방역소독 서비스", "pest_control"),
    "법정의무교육_서비스_구매요청서.xlsx": ("교육 서비스", "법정의무교육", "mandatory_education"),
    "보안경비_용역서비스_구매요청서.xlsx": ("건물 관리", "보안경비 용역 서비스", "security_guard"),
    "복합기_렌탈서비스_구매요청서.xlsx": ("건물 관리", "복합기 렌탈 서비스", "copier_rental"),
    "비데_렌탈서비스_구매요청서.xlsx": ("건물 관리", "비데 렌탈 서비스", "bidet_rental"),
    "안전관리_서비스_구매요청서.xlsx": ("건물 관리", "안전관리 서비스", "safety_management"),
    "어학교육_서비스_구매요청서.xlsx": ("교육 서비스", "어학교육", "language_education"),
    "전문교육_서비스_구매요청서.xlsx": ("교육 서비스", "전문교육", "professional_education"),
    "전자시장정보_구독서비스_구매요청서.xlsx": ("마케팅", "전자시장정보 구독", "market_data_subscription"),
    "조경관리_용역서비스_구매요청서.xlsx": ("건물 관리", "조경관리 용역 서비스", "landscaping"),
}


def _clean(val):
    if val is None:
        return ""
    s = str(val).strip()
    # 줄바꿈 제거 + 비BMP 문자(이모지) 제거
    s = s.replace("\n", " ").replace("\r", " ")
    return "".join(ch for ch in s if ord(ch) < 0x10000)


def parse_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if "구매요청서" not in wb.sheetnames:
        return None, None
    ws = wb["구매요청서"]
    sections = []
    items = []
    current_section = ""
    current_items = []

    for row_idx in range(1, ws.max_row + 1):
        row = [ws.cell(row_idx, c).value for c in range(1, min(ws.max_column + 1, 11))]
        first = _clean(row[0])

        # Section header
        new_section = None
        for cell in row:
            cv = _clean(cell)
            if cv.startswith("▶") or (cv.startswith("[") and "]" in cv):
                if "] " in cv:
                    new_section = cv.split("] ", 1)[-1].strip()
                else:
                    new_section = re.sub(r"^[▶\[\]\s]+", "", cv).strip()
                break

        if new_section:
            if current_section and current_items:
                sections.append((current_section, current_items[:]))
            current_section = new_section
            current_items = []
            continue

        if first.isdigit():
            name = _clean(row[2]) if len(row) > 2 else ""
            name = name[:60]
            req_str = _clean(row[4]) if len(row) > 4 else ""
            impact_str = _clean(row[5]) if len(row) > 5 else ""
            item = (int(first), name, "필수" in req_str, "상" in impact_str)
            items.append(item)
            current_items.append(item)

    if current_section and current_items:
        sections.append((current_section, current_items[:]))
    return items, sections


def generate():
    out = []
    out.append('"""구매요청서(PR) 스키마 — 신규DB Excel 16개 카테고리 기반 자동 생성')
    out.append("")
    out.append("필드 키: p1, p2, ... (RFP의 s1, s2, ...와 충돌 방지)")
    out.append('required: 필수 항목 또는 단가영향도 "상"인 항목')
    out.append('"""')
    out.append("")
    out.append("")
    out.append("PR_SCHEMAS = {")

    count = 0
    for fname in sorted(FILE_MAP.keys()):
        category, sub_cat, key = FILE_MAP[fname]
        path = os.path.join(EXCEL_DIR, fname)
        if not os.path.exists(path):
            continue

        items, sections = parse_excel(path)
        if not items:
            continue

        count += 1
        fields_str = ", ".join(f"p{no}:{name}" for no, name, _, _ in items)
        required_keys = [f"p{no}" for no, _, req, high in items if req or high]
        required_str = ",".join(required_keys)

        sec_lines = []
        sec_detail_parts = []
        for i, (sec_title, sec_items) in enumerate(sections, 1):
            names_short = ", ".join(name[:20] for _, name, _, _ in sec_items[:4])
            if len(sec_items) > 4:
                names_short += " 등"
            sec_lines.append(f"{i}. {sec_title} ({names_short})")
            field_keys = ",".join(f"p{no}" for no, _, _, _ in sec_items)
            sec_detail_parts.append(f"{sec_title}:{field_keys}")

        sections_str = "\\n".join(sec_lines)
        sections_detail = "|".join(sec_detail_parts)

        out.append(f'    "{key}": {{')
        out.append(f'        "label": "{sub_cat}",')
        out.append(f'        "category": "{category}",')
        out.append(f'        "sub_category": "{sub_cat}",')
        out.append(f'        "fields": "{fields_str}",')
        out.append(f'        "required": "{required_str}",')
        out.append(f'        "sections": "{sections_str}",')
        out.append(f'        "sections_detail": "{sections_detail}",')
        out.append("    },")

    # Generic fallback
    out.append('    "_generic": {')
    out.append('        "label": "일반 구매요청",')
    out.append('        "category": "",')
    out.append('        "sub_category": "",')
    out.append('        "fields": "p1:요청부서, p2:요청자, p3:연락처, p4:품목/서비스명, p5:구매목적, p6:수량/규모, p7:희망납기, p8:예산범위, p9:요구사양, p10:납품조건, p11:기타요구사항",')
    out.append('        "required": "p1,p2,p3,p4,p5,p6,p7",')
    out.append('        "sections": "1. 요청자 정보 (요청부서, 요청자, 연락처)\\n2. 구매 개요 (품목/서비스명, 구매목적, 수량/규모, 희망납기)\\n3. 상세 요건 (예산범위, 요구사양, 납품조건, 기타요구사항)",')
    out.append('        "sections_detail": "요청자 정보:p1,p2,p3|구매 개요:p4,p5,p6,p7|상세 요건:p8,p9,p10,p11",')
    out.append("    },")
    out.append("}")
    out.append("")
    out.append("")

    # TAXONOMY_TO_PR
    out.append("# 분류체계 → PR 카테고리 매핑 (키워드 기반)")
    out.append("TAXONOMY_TO_PR = {")
    out.append('    "건물 관리": {')
    out.append('        "공기청정기": "air_purifier_rental",')
    out.append('        "복합기": "copier_rental",')
    out.append('        "비데": "bidet_rental",')
    out.append('        "문서파기": "document_shredding",')
    out.append('        "물리보안": "physical_security",')
    out.append('        "CCTV": "physical_security",')
    out.append('        "방역": "pest_control",')
    out.append('        "소독": "pest_control",')
    out.append('        "보안경비": "security_guard",')
    out.append('        "경비": "security_guard",')
    out.append('        "안전관리": "safety_management",')
    out.append('        "안전": "safety_management",')
    out.append('        "승강기": "safety_management",')
    out.append('        "소방": "safety_management",')
    out.append('        "조경": "landscaping",')
    out.append('        "_default": "_generic",')
    out.append("    },")
    out.append('    "마케팅": {')
    out.append('        "디지털": "digital_ad",')
    out.append('        "광고": "digital_ad",')
    out.append('        "배너": "digital_ad",')
    out.append('        "바이럴": "viral_marketing",')
    out.append('        "인플루언서": "viral_marketing",')
    out.append('        "SNS": "viral_marketing",')
    out.append('        "블로그": "viral_marketing",')
    out.append('        "전자시장": "market_data_subscription",')
    out.append('        "GFK": "market_data_subscription",')
    out.append('        "NPD": "market_data_subscription",')
    out.append('        "_default": "_generic",')
    out.append("    },")
    out.append('    "교육 서비스": {')
    out.append('        "법정": "mandatory_education",')
    out.append('        "의무교육": "mandatory_education",')
    out.append('        "산업안전": "mandatory_education",')
    out.append('        "성희롱": "mandatory_education",')
    out.append('        "어학": "language_education",')
    out.append('        "영어": "language_education",')
    out.append('        "일본어": "language_education",')
    out.append('        "중국어": "language_education",')
    out.append('        "전문교육": "professional_education",')
    out.append('        "직무교육": "professional_education",')
    out.append('        "리더십": "professional_education",')
    out.append('        "_default": "_generic",')
    out.append("    },")
    out.append("}")
    out.append("")
    out.append("")

    # PR_PHASE_PROMPT
    out.append('PR_PHASE_PROMPT = """사용자 메시지에서 구매요청서 필드 값을 추출하여 JSON만 반환하세요. 설명 없이 JSON만.')
    out.append("")
    out.append("## 전체 필드 목록")
    out.append("{fields}")
    out.append("")
    out.append("## 섹션 구조")
    out.append("{sections}")
    out.append("")
    out.append("## 현재 상태")
    out.append("- 현재 phase: {phase}")
    out.append("- 이미 채워진 필드: {filled_keys}")
    out.append("- 주요 필수 필드: {required_keys}")
    out.append("")
    out.append("## 핵심 규칙")
    out.append("1. 대화 이력에서 AI가 마지막으로 요청한 섹션과 필드 목록을 정확히 확인하세요.")
    out.append("2. 사용자의 현재 메시지는 AI가 요청한 필드에 대한 답변입니다.")
    out.append("3. 사용자 입력을 AI가 요청한 필드 순서대로 1:1 매핑하세요.")
    out.append("   - 쉼표로 구분된 경우: 쉼표 기준으로 분리 후 순서대로 매핑")
    out.append("   - 쉼표가 부족한 경우: 의미 단위로 분리하세요.")
    out.append("   - 핵심: 요청 필드 N개에 대해 값 N개를 추출하여 순서대로 매핑하세요.")
    out.append("4. 하나의 필드에 여러 값을 합치지 마세요.")
    out.append("5. 아직 요청하지 않은 섹션의 필드에 값을 넣지 마세요.")
    out.append("6. 이미 채워진 필드는 다시 추출하지 마세요.")
    out.append("7. is_complete: 이미 채워진 필드 + 새로 추출한 필드로 주요 필수 필드가 모두 채워졌으면 true.")
    out.append("")
    out.append("## 대화 이력")
    out.append("{history}")
    out.append("")
    out.append("## 현재 사용자 메시지")
    out.append("{message}")
    out.append("")
    out.append('## 출력 형식')
    out.append('{{"pr_fields": {{"p1": "값", "p2": "값"}}, "is_complete": false}}"""')

    content = "\n".join(out)

    out_path = os.path.join(os.path.dirname(__file__), "..", "app", "constants", "pr_schemas.py")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(content)

    print(f"Generated {count} category schemas + generic + TAXONOMY_TO_PR + PR_PHASE_PROMPT")
    print(f"Output: {os.path.abspath(out_path)}")
    print(f"File size: {os.path.getsize(out_path):,} bytes")


if __name__ == "__main__":
    generate()
