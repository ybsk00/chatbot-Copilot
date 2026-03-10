"""
신규DB Excel → knowledge_chunks 청킹 + 임베딩 스크립트

Excel 구조:
  - Sheet 1: 구매요청서 (항목별 사양 + 설명 + 단가 영향도)
  - Sheet 2: 단가영향 요소 요약
  - Sheet 3: 시장단가 기준표 / 작성 가이드 / 법정의무 체크리스트
  + 바이럴마케팅_시장단가표.xlsx (시장단가 마스터 + 카테고리별 비교 + 소싱전략)

청킹 전략:
  1. 구매요청서 항목 → 2~3개씩 묶어 "Q: {카테고리} {항목}은? / A: 사양+설명+단가영향" 형태
  2. 단가영향 요소 → 영향도별 그룹핑 (상/중/하)
  3. 시장단가 → 항목별 가격+전략 청크 (역제안용 핵심 데이터)
  4. 소싱전략 → 예산규모별/협상전략별 청크

사용법:
  python -m scripts.ingest_excel
  python -m scripts.ingest_excel --dry-run   (DB 저장 없이 청크 확인)
"""
import sys
import os
import time
import argparse
import re

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from app.rag.embedder import embed_document
from app.db.supabase_client import get_client

EXCEL_DIR = os.path.join(
    os.path.dirname(__file__), "..", "..", "..", "신규db"
)

# 신규DB 파일 → (category, sub_cat) 매핑
FILE_CATEGORY_MAP = {
    "공기청정기_렌탈서비스_구매요청서.xlsx": ("건물 관리", "공기청정기 렌탈 서비스"),
    "디지털광고제작운영_구매요청서.xlsx": ("마케팅", "디지털 광고 제작 및 운영"),
    "문서파기_서비스_구매요청서.xlsx": ("건물 관리", "문서파기 서비스"),
    "물리보안_서비스_구매요청서.xlsx": ("건물 관리", "물리보안 서비스"),
    "바이럴마케팅_시장단가표.xlsx": ("마케팅", "바이럴 마케팅 대행"),
    "바이럴마케팅대행_구매요청서.xlsx": ("마케팅", "바이럴 마케팅 대행"),
    "방역소독_구독서비스_구매요청서.xlsx": ("건물 관리", "방역소독 서비스"),
    "법정의무교육_서비스_구매요청서.xlsx": ("교육 서비스", "법정의무교육"),
    "보안경비_용역서비스_구매요청서.xlsx": ("건물 관리", "보안경비 용역 서비스"),
    "복합기_렌탈서비스_구매요청서.xlsx": ("건물 관리", "복합기 렌탈 서비스"),
    "비데_렌탈서비스_구매요청서.xlsx": ("건물 관리", "비데 렌탈 서비스"),
    "안전관리_서비스_구매요청서.xlsx": ("건물 관리", "안전관리 서비스"),
    "어학교육_서비스_구매요청서.xlsx": ("교육 서비스", "어학교육"),
    "전문교육_서비스_구매요청서.xlsx": ("교육 서비스", "전문교육"),
    "전자시장정보_구독서비스_구매요청서.xlsx": ("마케팅", "전자시장정보 구독"),
    "조경관리_용역서비스_구매요청서.xlsx": ("건물 관리", "조경관리 용역 서비스"),
}


def _clean(val):
    """셀 값 정리"""
    if val is None:
        return ""
    s = str(val).strip()
    # 이모지 제거 (카테고리 헤더용)
    s = re.sub(r'[^\w\s가-힣a-zA-Z0-9.,()/%·\-+~×&:;★☆\[\]|→↑↓<>{}₩$€¥]', '', s)
    return s.strip()


def _is_section_header(row_values):
    """섹션 헤더인지 판별 (▶ 로 시작하는 행)"""
    first = _clean(row_values[0]) if row_values else ""
    return first.startswith("▶") or first.startswith("[")


def chunk_purchase_request(ws, category, sub_cat, doc_name):
    """구매요청서 시트 → 항목별 청크 생성

    2~3개 항목을 묶어서 하나의 청크로:
    Q: {sub_cat} 구매 시 {항목명}은 어떻게 해야 하나요?
    [{항목명}] {설명}
    - 필수/옵션: {필수}
    - 단가 영향도: {상/중/하}
    - 입력 예시: {예시}
    """
    chunks = []
    items = []
    current_section = ""

    for row_idx in range(1, ws.max_row + 1):
        row = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]

        # 헤더/빈 행 스킵
        first_val = _clean(row[0]) if row[0] else ""
        if not first_val or first_val in ("No.", "요청일자", "예산코드", ""):
            # 섹션 헤더 감지
            for cell in row:
                cv = _clean(cell) if cell else ""
                if cv.startswith("▶") or cv.startswith("["):
                    current_section = re.sub(r'^[\▶\[\]\s]+', '', cv).strip()
                    break
            continue

        # 숫자로 시작하는 행 = 항목
        if first_val.isdigit():
            cols = [_clean(c) for c in row]
            item = {
                "no": cols[0] if len(cols) > 0 else "",
                "section": current_section,
                "classification": cols[1] if len(cols) > 1 else "",
                "name": cols[2] if len(cols) > 2 else "",
                "description": cols[3] if len(cols) > 3 else "",
                "required": cols[4] if len(cols) > 4 else "",
                "price_impact": cols[5] if len(cols) > 5 else "",
                "example": cols[6] if len(cols) > 6 else "",
                "note": cols[7] if len(cols) > 7 else "",
                "extra": cols[8] if len(cols) > 8 else "",
            }
            if item["name"]:
                items.append(item)

    # 2~3개씩 묶어서 청크 생성
    group_size = 3
    for i in range(0, len(items), group_size):
        group = items[i:i + group_size]
        item_names = ", ".join(it["name"] for it in group)

        # Q&A 형태 구성
        question = f"{sub_cat} 구매 시 {item_names}은 어떻게 설정하나요?"

        body_parts = []
        for it in group:
            parts = [f"[{it['name']}]"]
            if it["description"]:
                parts.append(it["description"])
            meta_parts = []
            if it["required"]:
                meta_parts.append(f"필수여부: {it['required']}")
            if it["price_impact"]:
                meta_parts.append(f"단가 영향도: {it['price_impact']}")
            if meta_parts:
                parts.append(" | ".join(meta_parts))
            if it["example"]:
                parts.append(f"입력 예시: {it['example']}")
            if it["note"]:
                parts.append(f"참고: {it['note']}")
            body_parts.append("\n".join(parts))

        content = f"Q: {question}\n\n" + "\n\n".join(body_parts)

        chunks.append({
            "category": category,
            "sub_cat": sub_cat,
            "doc_name": doc_name,
            "chunk_index": len(chunks),
            "content": content,
            "metadata": {
                "type": "purchase_request",
                "items": [it["name"] for it in group],
                "section": group[0].get("section", ""),
                "price_impacts": [it["price_impact"] for it in group],
            },
        })

    return chunks


def chunk_price_impact(ws, category, sub_cat, doc_name):
    """단가영향 요소 요약 시트 → 항목별 상세 청크

    열 구조: No. | 분류 | 항목명 | 단가 영향 요인 상세 | 조건별 단가 변동 범위 | 소싱 협상 포인트 | 비고
    """
    chunks = []
    current_section = ""  # 단가영향 상/중/하 섹션
    items = []

    for row_idx in range(1, ws.max_row + 1):
        row = [_clean(ws.cell(row_idx, c).value) for c in range(1, ws.max_column + 1)]
        first = row[0] if row else ""

        # 섹션 헤더 감지 (▶ 단가영향 상 — TCO 핵심 결정 요소)
        full_row = " ".join(r for r in row if r)
        if "단가영향" in full_row and ("상" in full_row or "중" in full_row or "하" in full_row) and "항목" not in full_row:
            # 이전 섹션 저장
            if current_section and items:
                _save_price_impact_chunk(chunks, current_section, items, category, sub_cat, doc_name)
                items = []
            if "상" in full_row and ("10%" in full_row or "핵심" in full_row or "TCO" in full_row):
                current_section = "단가영향 상 (10% 이상, TCO 핵심)"
            elif "중" in full_row:
                current_section = "단가영향 중 (5~10%)"
            elif "하" in full_row:
                current_section = "단가영향 하 (5% 미만)"
            continue

        # 데이터 행: No.가 숫자인 행
        if not first or not first.isdigit():
            continue

        classification = row[1] if len(row) > 1 else ""  # 분류 (공통/제품/필터/정산)
        item_name = row[2] if len(row) > 2 else ""        # 항목명
        detail = row[3] if len(row) > 3 else ""            # 단가 영향 요인 상세
        price_range = row[4] if len(row) > 4 else ""       # 조건별 단가 변동 범위
        negotiation = row[5] if len(row) > 5 else ""       # 소싱 협상 포인트
        note = row[6] if len(row) > 6 else ""              # 비고

        if not item_name:
            continue

        items.append({
            "classification": classification,
            "item_name": item_name,
            "detail": detail,
            "price_range": price_range,
            "negotiation": negotiation,
            "note": note,
        })

    # 마지막 섹션
    if current_section and items:
        _save_price_impact_chunk(chunks, current_section, items, category, sub_cat, doc_name)

    # 섹션 헤더가 없는 경우 전체를 하나로
    if not chunks and items:
        _save_price_impact_chunk(chunks, "단가영향 요소", items, category, sub_cat, doc_name)

    return chunks


def _save_price_impact_chunk(chunks, section, items, category, sub_cat, doc_name):
    """단가영향 항목 2~3개씩 묶어서 청크 생성"""
    group_size = 2
    for i in range(0, len(items), group_size):
        group = items[i:i + group_size]
        item_names = ", ".join(it["item_name"] for it in group)

        question = f"{sub_cat} 단가 영향 요소: {item_names}의 단가 변동과 협상 전략은?"

        body_parts = []
        for it in group:
            lines = [f"[{it['classification']} - {it['item_name']}] ({section})"]
            if it["detail"]:
                lines.append(f"영향 요인: {it['detail']}")
            if it["price_range"]:
                lines.append(f"단가 변동 범위: {it['price_range']}")
            if it["negotiation"]:
                lines.append(f"소싱 협상 포인트: {it['negotiation']}")
            if it["note"]:
                lines.append(f"비고: {it['note']}")
            body_parts.append("\n".join(lines))

        content = f"Q: {question}\n\n" + "\n\n".join(body_parts)

        chunks.append({
            "category": category,
            "sub_cat": sub_cat,
            "doc_name": doc_name,
            "chunk_index": len(chunks),
            "content": content,
            "metadata": {
                "type": "price_impact",
                "section": section,
                "items": [it["item_name"] for it in group],
            },
        })


def chunk_market_price(ws, category, sub_cat, doc_name):
    """시장단가 마스터 시트 → 항목별 가격 청크 (역제안 핵심 데이터)"""
    chunks = []
    items = []
    current_section = ""

    for row_idx in range(1, ws.max_row + 1):
        row = [_clean(ws.cell(row_idx, c).value) for c in range(1, ws.max_column + 1)]
        first = row[0] if row else ""

        # 섹션 헤더
        if first.startswith("▶") or first.startswith("["):
            current_section = re.sub(r'^[\▶\[\]\s]+', '', first).strip()
            continue

        if not first or not first.isdigit():
            continue

        item = {
            "no": first,
            "section": current_section,
            "service": row[1] if len(row) > 1 else "",
            "detail": row[2] if len(row) > 2 else "",
            "unit": row[3] if len(row) > 3 else "",
            "min_price": row[4] if len(row) > 4 else "",
            "avg_price": row[5] if len(row) > 5 else "",
            "max_price": row[6] if len(row) > 6 else "",
            "quality": row[7] if len(row) > 7 else "",
            "includes": row[8] if len(row) > 8 else "",
            "factors": row[9] if len(row) > 9 else "",
            "tips": row[10] if len(row) > 10 else "",
            "negotiability": row[11] if len(row) > 11 else "",
        }
        if item["service"]:
            items.append(item)

    # 2개씩 묶어서 비교 가능한 청크 생성
    group_size = 2
    for i in range(0, len(items), group_size):
        group = items[i:i + group_size]
        names = " vs ".join(f"{it['service']}({it['detail']})" for it in group)

        question = f"{sub_cat} 시장단가: {names} 비용은 얼마인가요?"

        body_parts = []
        for it in group:
            lines = [f"[{it['service']} - {it['detail']}]"]
            if it["unit"]:
                lines.append(f"과금 단위: {it['unit']}")
            if it["min_price"] and it["avg_price"] and it["max_price"]:
                lines.append(f"시장가: 최저 {it['min_price']}원 / 평균 {it['avg_price']}원 / 최고 {it['max_price']}원")
            if it["quality"]:
                lines.append(f"품질 레벨: {it['quality']}")
            if it["includes"]:
                lines.append(f"포함 내용: {it['includes']}")
            if it["factors"]:
                lines.append(f"단가 영향 요인: {it['factors']}")
            if it["tips"]:
                lines.append(f"구매 전략: {it['tips']}")
            if it["negotiability"]:
                lines.append(f"협상 가능성: {it['negotiability']}")
            body_parts.append("\n".join(lines))

        content = f"Q: {question}\n\n" + "\n\n".join(body_parts)

        chunks.append({
            "category": category,
            "sub_cat": sub_cat,
            "doc_name": doc_name,
            "chunk_index": len(chunks),
            "content": content,
            "metadata": {
                "type": "market_price",
                "items": [f"{it['service']}({it['detail']})" for it in group],
                "price_range": {
                    "min": group[0].get("min_price", ""),
                    "max": group[0].get("max_price", ""),
                },
            },
        })

    return chunks


def chunk_sourcing_strategy(ws, category, sub_cat, doc_name):
    """소싱전략 가이드 시트 → 전략별 청크"""
    chunks = []
    current_section = ""
    section_items = []

    for row_idx in range(1, ws.max_row + 1):
        row = [_clean(ws.cell(row_idx, c).value) for c in range(1, ws.max_column + 1)]
        first = row[0] if row else ""

        # 숫자+이모지 섹션 헤더 (1️⃣, 2️⃣ 등)
        if any(marker in first for marker in ["예산", "협상", "공급", "발주", "전략", "체크"]):
            # 이전 섹션 저장
            if current_section and section_items:
                _save_strategy_chunk(chunks, current_section, section_items,
                                     category, sub_cat, doc_name)
            current_section = first
            section_items = []
            continue

        # 데이터 행
        if first and first not in ("구분", "카테고리", "No."):
            entry_parts = [first]
            for col_val in row[1:]:
                if col_val:
                    entry_parts.append(col_val)
            if len(entry_parts) > 1:
                section_items.append(" | ".join(entry_parts))

    # 마지막 섹션
    if current_section and section_items:
        _save_strategy_chunk(chunks, current_section, section_items,
                             category, sub_cat, doc_name)

    return chunks


def _save_strategy_chunk(chunks, section, items, category, sub_cat, doc_name):
    question = f"{sub_cat} 소싱 전략: {section}?"
    content = f"Q: {question}\n\n" + "\n".join(f"- {item}" for item in items)
    chunks.append({
        "category": category,
        "sub_cat": sub_cat,
        "doc_name": doc_name,
        "chunk_index": len(chunks),
        "content": content,
        "metadata": {
            "type": "sourcing_strategy",
            "section": section,
        },
    })


def chunk_guide_checklist(ws, category, sub_cat, doc_name):
    """작성 가이드 / 법정의무 체크리스트 / 기준표 시트 → 섹션별 청크

    섹션 구분 기준:
    - 빈 행 뒤에 나오는 텍스트 행 = 새 섹션 헤더
    - 첫 번째 열에만 값이 있고 나머지가 비어있는 행 = 섹션 헤더
    - 시트 타이틀 행(1~3행)은 스킵
    """
    chunks = []
    sections = []  # [(section_title, [rows...])]
    sheet_title = ""  # 시트 상단 타이틀 (폴백용)
    current_title = ""
    current_rows = []
    prev_blank = False

    for row_idx in range(1, ws.max_row + 1):
        raw = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
        cleaned = [_clean(v) for v in raw]
        non_empty = [v for v in cleaned if v]

        # 빈 행
        if not non_empty:
            prev_blank = True
            continue

        # 시트 타이틀 (첫 3행 중 1열만 채워진 행) → 폴백 타이틀로 저장
        if row_idx <= 3 and len(non_empty) == 1:
            if not sheet_title:
                sheet_title = non_empty[0].strip()[:60]
            prev_blank = True
            continue

        # 섹션 헤더 판별: (빈 행 뒤 + 1열만 값 있음) 또는 (빈 행 뒤 + No./구분 등 컬럼 헤더)
        is_header = False
        first_val = cleaned[0]

        # 조건 1: 빈 행 뒤에 나온 행 중, 1~2열만 채워진 경우 → 섹션 헤더
        if prev_blank and len(non_empty) <= 2 and not first_val.isdigit():
            is_header = True
        # 조건 2: "No." 또는 "구분" 등 테이블 컬럼 헤더 → 섹션 데이터 헤더 (스킵)
        if first_val in ("No.", "구분", "No"):
            prev_blank = False
            continue

        if is_header:
            # 이전 섹션 저장 (타이틀 없는 초기 데이터는 시트 타이틀 사용)
            if current_rows:
                title = current_title or sheet_title or f"{sub_cat} 가이드"
                sections.append((title, current_rows))
            current_title = " ".join(non_empty)
            current_rows = []
            prev_blank = False
            continue

        # 데이터 행
        row_text = " | ".join(v for v in cleaned if v)
        if row_text:
            current_rows.append(row_text)
        prev_blank = False

    # 마지막 섹션
    if current_rows:
        title = current_title or sheet_title or f"{sub_cat} 가이드"
        sections.append((title, current_rows))

    # 섹션이 없으면 전체를 하나로
    if not sections and current_rows:
        sections = [(f"{sub_cat} 가이드", current_rows)]

    # 섹션별 청크 생성 (너무 큰 섹션은 분할)
    max_rows_per_chunk = 12
    for title, rows in sections:
        if len(rows) <= max_rows_per_chunk:
            # 한 청크로
            question = f"{sub_cat} {title}"
            content = f"Q: {question}\n\n" + "\n".join(f"- {r}" for r in rows)
            if len(content) < 30:
                continue
            chunks.append({
                "category": category,
                "sub_cat": sub_cat,
                "doc_name": doc_name,
                "chunk_index": len(chunks),
                "content": content,
                "metadata": {
                    "type": "guide_checklist",
                    "section": title,
                    "sheet_name": ws.title,
                },
            })
        else:
            # 큰 섹션은 분할
            for i in range(0, len(rows), max_rows_per_chunk):
                group = rows[i:i + max_rows_per_chunk]
                part = f" (Part {i // max_rows_per_chunk + 1})" if len(rows) > max_rows_per_chunk else ""
                question = f"{sub_cat} {title}{part}"
                content = f"Q: {question}\n\n" + "\n".join(f"- {r}" for r in group)
                if len(content) < 30:
                    continue
                chunks.append({
                    "category": category,
                    "sub_cat": sub_cat,
                    "doc_name": doc_name,
                    "chunk_index": len(chunks),
                    "content": content,
                    "metadata": {
                        "type": "guide_checklist",
                        "section": title,
                        "sheet_name": ws.title,
                    },
                })

    return chunks


def chunk_category_comparison(ws, category, sub_cat, doc_name):
    """카테고리별 단가 비교 시트 → 비교표 청크"""
    chunks = []
    rows = []

    for row_idx in range(1, ws.max_row + 1):
        row = [_clean(ws.cell(row_idx, c).value) for c in range(1, ws.max_column + 1)]
        first = row[0] if row else ""
        if not first or first in ("카테고리",):
            continue
        rows.append(row)

    if not rows:
        return chunks

    # 전체를 하나의 비교 청크로
    body_parts = []
    for row in rows:
        cat_name = row[0] if len(row) > 0 else ""
        service = row[1] if len(row) > 1 else ""
        min_p = row[2] if len(row) > 2 else ""
        avg_p = row[3] if len(row) > 3 else ""
        max_p = row[4] if len(row) > 4 else ""
        unit = row[5] if len(row) > 5 else ""
        roi = row[6] if len(row) > 6 else ""
        target = row[7] if len(row) > 7 else ""

        line = f"- {cat_name} ({service}): {min_p}~{max_p}원 (평균 {avg_p}원, {unit})"
        if roi:
            line += f" ROI: {roi}"
        if target:
            line += f" 추천: {target}"
        body_parts.append(line)

    question = f"{sub_cat} 카테고리별 단가 비교 — 어떤 채널이 가장 효율적인가요?"
    content = f"Q: {question}\n\n" + "\n".join(body_parts)

    chunks.append({
        "category": category,
        "sub_cat": sub_cat,
        "doc_name": doc_name,
        "chunk_index": 0,
        "content": content,
        "metadata": {
            "type": "category_comparison",
        },
    })

    return chunks


def process_excel(filepath, dry_run=False):
    """단일 Excel 파일 처리"""
    filename = os.path.basename(filepath)
    if filename not in FILE_CATEGORY_MAP:
        print(f"  [!] 매핑 없음: {filename}")
        return {"file": filename, "saved": 0, "total": 0}

    category, sub_cat = FILE_CATEGORY_MAP[filename]
    wb = openpyxl.load_workbook(filepath)
    all_chunks = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sn_lower = sheet_name.lower()

        if "구매요청서" in sheet_name:
            all_chunks.extend(chunk_purchase_request(ws, category, sub_cat, filename))
        elif "단가영향" in sheet_name:
            all_chunks.extend(chunk_price_impact(ws, category, sub_cat, filename))
        elif "시장단가 마스터" in sheet_name or "시장단가 기준표" in sheet_name:
            all_chunks.extend(chunk_market_price(ws, category, sub_cat, filename))
        elif "카테고리별" in sheet_name:
            all_chunks.extend(chunk_category_comparison(ws, category, sub_cat, filename))
        elif "소싱전략" in sheet_name:
            all_chunks.extend(chunk_sourcing_strategy(ws, category, sub_cat, filename))
        elif any(kw in sheet_name for kw in ("가이드", "코드표", "체크리스트", "기준표", "평가")):
            all_chunks.extend(chunk_guide_checklist(ws, category, sub_cat, filename))
        elif "단가영향요소" in sheet_name.replace(" ", ""):
            all_chunks.extend(chunk_price_impact(ws, category, sub_cat, filename))
        else:
            print(f"    [i] 미처리 시트: {sheet_name}")

    # chunk_index 재정렬
    for idx, chunk in enumerate(all_chunks):
        chunk["chunk_index"] = idx

    if dry_run:
        for chunk in all_chunks:
            print(f"\n--- [{chunk['metadata']['type']}] chunk {chunk['chunk_index']} ---")
            print(chunk["content"][:300])
            print(f"  metadata: {chunk['metadata']}")
        return {"file": filename, "saved": 0, "total": len(all_chunks)}

    # 임베딩 + DB 저장
    supabase = get_client()
    saved = 0
    for chunk in all_chunks:
        try:
            embedding = embed_document(chunk["content"])
            supabase.table("knowledge_chunks").insert({
                "category": chunk["category"],
                "sub_cat": chunk["sub_cat"],
                "doc_name": chunk["doc_name"],
                "chunk_index": chunk["chunk_index"],
                "content": chunk["content"],
                "metadata": chunk["metadata"],
                "embedding": embedding,
            }).execute()
            saved += 1
            time.sleep(0.15)  # API rate limit
        except Exception as e:
            print(f"    [X] 청크 {chunk['chunk_index']} 저장 실패: {e}")

    return {"file": filename, "saved": saved, "total": len(all_chunks)}


def main():
    # Windows cp949 인코딩 문제 방지
    if sys.platform == "win32":
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace", line_buffering=True)

    parser = argparse.ArgumentParser(description="신규DB Excel → knowledge_chunks 청킹")
    parser.add_argument("--dry-run", action="store_true", help="DB 저장 없이 청크 미리보기")
    parser.add_argument("--file", default=None, help="특정 파일만 처리")
    args = parser.parse_args()

    excel_dir = os.path.normpath(EXCEL_DIR)
    if not os.path.isdir(excel_dir):
        print(f"[X] 디렉토리 없음: {excel_dir}")
        sys.exit(1)

    files = sorted(f for f in os.listdir(excel_dir) if f.endswith(".xlsx") and not f.startswith("~$"))
    if args.file:
        files = [f for f in files if args.file in f]

    print(f"\n[*] 신규DB Excel 청킹 시작 ({len(files)}개 파일)")
    print(f"    경로: {excel_dir}")
    if args.dry_run:
        print("    [DRY RUN] DB 저장 안 함\n")

    total_saved = 0
    total_chunks = 0

    for f in files:
        filepath = os.path.join(excel_dir, f)
        print(f"\n[>] {f}")
        result = process_excel(filepath, dry_run=args.dry_run)
        print(f"    -> {result['saved']}/{result['total']} 청크 저장")
        total_saved += result["saved"]
        total_chunks += result["total"]

    print(f"\n{'[DRY RUN]' if args.dry_run else '[DONE]'} 완료! 총 {total_saved}/{total_chunks} 청크 저장\n")


if __name__ == "__main__":
    main()
