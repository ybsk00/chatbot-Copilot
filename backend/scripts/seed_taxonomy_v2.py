"""차세대 품목체계 시드 — Excel → taxonomy_v2 테이블 (L1/L2/L3 3레벨)"""
import sys, os, pathlib
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from dotenv import load_dotenv
load_dotenv()

import openpyxl
from app.db.supabase_client import get_client


# ═══════════════════════════════════════════
# L2 → PR 템플릿 그룹 매핑 (59개 L2 → 20개 PR 그룹)
# ═══════════════════════════════════════════
L2_TO_PR_TEMPLATE = {
    # 1. 물품구매 — 물품/소모품 구매
    "L2-0101": "goods_purchase",       # 사무용품·소모품
    "L2-1002": "goods_purchase",       # 공장용 소모품

    # 2. 렌탈·리스 — 기기/장비 렌탈·리스
    "L2-0104": "rental_lease",         # 사무환경 관리
    "L2-0303": "rental_lease",         # 비품 렌탈·구매
    "L2-0401": "rental_lease",         # 법인차량

    # 3. 인쇄·제작 — 인쇄물/디자인/영상 제작
    "L2-0102": "print_production",     # 인쇄 서비스
    "L2-0703": "print_production",     # 콘텐츠·크리에이티브 제작
    "L2-0705": "print_production",     # 판촉·판매촉진

    # 4. 용역·아웃소싱 — 인력 투입형 용역
    "L2-0302": "service_outsourcing",  # 건물관리 용역 (FM)
    "L2-0203": "service_outsourcing",  # 채용·인력
    "L2-0604": "service_outsourcing",  # 기타 전문용역 (컨택센터)
    "L2-1004": "service_outsourcing",  # 생산 운영 서비스

    # 5. 전문 컨설팅 — 전문가 자문/컨설팅
    "L2-0601": "consulting",           # 법무·소송
    "L2-0602": "consulting",           # 경영컨설팅
    "L2-0603": "consulting",           # 회계·세무·급여대행

    # 6. 교육·훈련 — 교육/연수/학습
    "L2-0202": "education_training",   # 교육·훈련

    # 7. 복지 서비스 — 임직원 복지/편의
    "L2-0201": "welfare_service",      # 임직원 복지 서비스
    "L2-0204": "welfare_service",      # 글로벌 HR 서비스

    # 8. 시설 공사 — 건설/인테리어/설비 공사
    "L2-0301": "construction",         # 시설공사
    "L2-1003": "construction",         # 시설·설비공사·유지보수

    # 9. 보험 — 모든 보험 유형
    "L2-0501": "insurance",            # 임직원 단체 보험
    "L2-0502": "insurance",            # 자산·시설 보험
    "L2-0503": "insurance",            # 배상책임 보험
    "L2-0504": "insurance",            # 운송·물류 보험
    "L2-0505": "insurance",            # 법정·의무 보험
    "L2-0506": "insurance",            # 보증 보험
    "L2-0507": "insurance",            # 차량·출장 보험
    "L2-0508": "insurance",            # 기타 보험

    # 10. 광고·매체 — 광고 집행/매체 구매
    "L2-0701": "advertising_media",    # 일반 광고·매체
    "L2-0702": "advertising_media",    # 디지털 마케팅
    "L2-0706": "advertising_media",    # 홍보·PR
    "L2-0707": "advertising_media",    # 리서치·Mar Tech

    # 11. 이벤트·행사 — 이벤트/전시/행사 운영
    "L2-0704": "event_exhibition",     # 이벤트·전시·행사

    # 12. 출장 서비스 — 출장/여행/이동
    "L2-0402": "business_travel",      # 출장 서비스

    # 13. IT HW — 서버/PC/네트워크 장비
    "L2-0801": "it_hardware",          # 서버·스토리지
    "L2-0802": "it_hardware",          # 엔드포인트·주변기기
    "L2-0803": "it_hardware",          # 네트워크·통신
    "L2-0807": "it_hardware",          # 데이터센터/인프라

    # 14. IT SW·클라우드 — 소프트웨어/클라우드/보안
    "L2-0804": "it_software_cloud",    # 소프트웨어
    "L2-0805": "it_software_cloud",    # 클라우드·보안

    # 15. IT 개발·운영 — SI/개발/ITO
    "L2-0806": "it_development",       # IT 서비스

    # 16. 물류 운송 — 운송/배송
    "L2-0901": "logistics_transport",  # 국내 육운
    "L2-0902": "logistics_transport",  # 국제 해운
    "L2-0903": "logistics_transport",  # 국제 항공·특송
    "L2-0907": "logistics_transport",  # 물류 부가 서비스

    # 17. 물류 창고·풀필먼트 — 창고/보관/풀필먼트
    "L2-0904": "logistics_warehouse",  # 창고·보관
    "L2-0905": "logistics_warehouse",  # 풀필먼트
    "L2-0906": "logistics_warehouse",  # 물류 IT/장비

    # 18. 연구 장비·소재 — 연구용 장비/부품/소재
    "L2-1101": "rd_service",           # 개발 용역 서비스
    "L2-1102": "rd_service",           # 개발용 소재·부품·시제품
    "L2-1103": "rd_service",           # 연구 장비·계측기
    "L2-1104": "rd_service",           # 연구개발 SW·IT 인프라

    # 19. 인증·시험·IP — 인증/시험/특허/로열티
    "L2-1105": "certification_test",   # 시험·인증·분석 서비스
    "L2-1106": "certification_test",   # 지식재산·로열티

    # 20. 생산·안전·폐기 — 생산설비/안전/폐기물
    "L2-1001": "production_safety",    # 생산설비·장비
    "L2-1005": "production_safety",    # 안전·보건·EHS
    "L2-1006": "production_safety",    # 불용자산·폐기물 관리

    # 21. 유틸리티·에너지 — 전기/가스/수도/에너지
    "L2-0304": "utility_energy",       # 유틸리티·에너지관리

    # 22. 사무보조 서비스 — 택배/번역/데이터
    "L2-0103": "office_support",       # 사무보조 서비스
}

# L2 → RFP 유형 매핑
L2_TO_RFP_TYPE = {
    # 물품 구매
    "L2-0101": "purchase", "L2-1002": "purchase",
    # 렌탈·리스
    "L2-0104": "rental", "L2-0303": "rental", "L2-0401": "rental",
    # 용역
    "L2-0102": "service_contract", "L2-0302": "service_contract",
    "L2-0203": "service_contract", "L2-0604": "service_contract",
    "L2-1004": "service_contract", "L2-0103": "service_contract",
    # 컨설팅
    "L2-0601": "consulting", "L2-0602": "consulting", "L2-0603": "consulting",
    # 공사
    "L2-0301": "construction", "L2-1003": "construction",
    # 서비스
    "L2-0201": "service", "L2-0202": "service", "L2-0204": "service",
    "L2-0402": "service", "L2-0304": "service",
    "L2-0501": "service", "L2-0502": "service", "L2-0503": "service",
    "L2-0504": "service", "L2-0505": "service", "L2-0506": "service",
    "L2-0507": "service", "L2-0508": "service",
    # 구매+유지보수
    "L2-0801": "purchase_maintenance", "L2-0802": "purchase_maintenance",
    "L2-0803": "purchase_maintenance", "L2-0807": "purchase_maintenance",
    "L2-1001": "purchase_maintenance", "L2-1103": "purchase_maintenance",
    # SW/클라우드 → 구매+리스
    "L2-0804": "purchase_lease", "L2-0805": "purchase_lease",
    # 마케팅 → 서비스
    "L2-0701": "service_contract", "L2-0702": "service_contract",
    "L2-0703": "service_contract", "L2-0704": "service_contract",
    "L2-0705": "service_contract", "L2-0706": "service_contract",
    "L2-0707": "service_contract",
    # IT 개발 → 서비스
    "L2-0806": "service_contract",
    # 물류 → 서비스
    "L2-0901": "service_contract", "L2-0902": "service_contract",
    "L2-0903": "service_contract", "L2-0904": "service_contract",
    "L2-0905": "service_contract", "L2-0906": "purchase_maintenance",
    "L2-0907": "service_contract",
    # R&D
    "L2-1101": "service_contract", "L2-1102": "purchase",
    "L2-1104": "purchase_lease",
    "L2-1105": "service_contract", "L2-1106": "service_contract",
    # 생산
    "L2-1005": "service_contract", "L2-1006": "service_contract",
}


def parse_excel(filepath: str) -> list[dict]:
    """Excel 파싱 → L1/L2/L3 레코드 리스트"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[0]

    records = []
    current_l1_code = None
    current_l1_name = None
    current_l2_code = None
    current_l2_name = None
    seen_l1 = set()
    seen_l2 = set()
    seen_l3 = set()

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        (l1_code, l1_name, l2_code, l2_name, l3_code, l3_name,
         cost_category, capex_opex, suppliers, strategy, description, keywords) = row[:12]

        # L1 갱신
        if l1_code:
            current_l1_code = str(l1_code).strip()
            current_l1_name = str(l1_name).strip()
            if current_l1_code not in seen_l1:
                seen_l1.add(current_l1_code)
                records.append({
                    "code": current_l1_code,
                    "level": 1,
                    "name": current_l1_name,
                    "parent_code": None,
                })

        # L2 갱신
        if l2_code:
            current_l2_code = str(l2_code).strip()
            current_l2_name = str(l2_name).strip()
            if current_l2_code not in seen_l2:
                seen_l2.add(current_l2_code)
                pr_key = L2_TO_PR_TEMPLATE.get(current_l2_code, "_generic")
                rfp_type = L2_TO_RFP_TYPE.get(current_l2_code, "service")
                records.append({
                    "code": current_l2_code,
                    "level": 2,
                    "name": current_l2_name,
                    "parent_code": current_l1_code,
                    "pr_template_key": pr_key,
                    "rfp_type": rfp_type,
                })

        # L3
        if l3_code:
            l3_code_str = str(l3_code).strip()
            l3_name_str = str(l3_name).strip()

            # Excel 코드 중복 보정 (L2-1106 지식재산 코드가 L2-1105와 충돌)
            if l3_code_str in seen_l3:
                # 부모 L2 코드 기반으로 새 코드 생성
                l2_num = current_l2_code.replace("L2-", "")
                suffix = 1
                while f"L3-{l2_num}{suffix:02d}" in seen_l3:
                    suffix += 1
                old_code = l3_code_str
                l3_code_str = f"L3-{l2_num}{suffix:02d}"
                print(f"  [보정] {old_code} → {l3_code_str} ({l3_name_str})")
            seen_l3.add(l3_code_str)

            # 키워드 파싱
            kw_list = []
            if keywords:
                kw_list = [k.strip() for k in str(keywords).split(",") if k.strip()]

            # 공급사 파싱
            sup_list = []
            if suppliers:
                sup_list = [s.strip() for s in str(suppliers).split(",") if s.strip()]

            # CAPEX/OPEX
            expense = "OPEX"
            if capex_opex:
                co = str(capex_opex).strip().upper()
                if "CAPEX" in co and "OPEX" in co:
                    expense = "BOTH"
                elif "CAPEX" in co:
                    expense = "CAPEX"

            # 비용 분류
            cost_cat = str(cost_category).strip() if cost_category else None

            # PR 템플릿/RFP 유형은 부모 L2에서 상속
            pr_key = L2_TO_PR_TEMPLATE.get(current_l2_code, "_generic")
            rfp_type = L2_TO_RFP_TYPE.get(current_l2_code, "service")

            records.append({
                "code": l3_code_str,
                "level": 3,
                "name": l3_name_str,
                "parent_code": current_l2_code,
                "description": str(description).strip() if description else None,
                "keywords": kw_list,
                "purchase_strategy": str(strategy).strip() if strategy else None,
                "expense_type": expense,
                "cost_category": cost_cat,
                "suppliers": sup_list,
                "pr_template_key": pr_key,
                "rfp_type": rfp_type,
            })

    return records


def find_excel() -> pathlib.Path:
    """Windows 한글 경로 호환 — pathlib으로 Excel 파일 탐색"""
    # scripts/ → backend/ → ip-assist/ → 업무마켓챗봇/
    for depth in range(2, 5):
        base = pathlib.Path(__file__).resolve()
        for _ in range(depth):
            base = base.parent
        for f in base.glob("*.xlsx"):
            if "0317" in f.name:
                return f
    return None


def run():
    excel_path = find_excel()
    if not excel_path or not excel_path.exists():
        print("Excel 파일을 찾을 수 없습니다 (차세대_품목체계_설명_0317.xlsx)")
        return
    print(f"Excel 파일: {excel_path.name}")

    records = parse_excel(excel_path)
    print(f"파싱 완료: L1={sum(1 for r in records if r['level']==1)}개, "
          f"L2={sum(1 for r in records if r['level']==2)}개, "
          f"L3={sum(1 for r in records if r['level']==3)}개, "
          f"총 {len(records)}개")

    supabase = get_client()

    # 기존 데이터 삭제
    supabase.table("taxonomy_v2").delete().neq("id", 0).execute()
    print("기존 taxonomy_v2 데이터 삭제 완료")

    # 배치 삽입 (50개씩)
    batch_size = 50
    inserted = 0
    for i in range(0, len(records), batch_size):
        batch = records[i:i + batch_size]
        # None 값 필터링 (Supabase는 None 키를 보내면 안 됨)
        clean_batch = []
        for r in batch:
            row = {"code": r["code"], "level": r["level"], "name": r["name"], "is_active": True}
            if r.get("parent_code"):
                row["parent_code"] = r["parent_code"]
            if r.get("description"):
                row["description"] = r["description"]
            if r.get("keywords"):
                row["keywords"] = r["keywords"]
            if r.get("purchase_strategy"):
                row["purchase_strategy"] = r["purchase_strategy"]
            if r.get("expense_type"):
                row["expense_type"] = r["expense_type"]
            if r.get("cost_category"):
                row["cost_category"] = r["cost_category"]
            if r.get("suppliers"):
                row["suppliers"] = r["suppliers"]
            if r.get("pr_template_key"):
                row["pr_template_key"] = r["pr_template_key"]
            if r.get("rfp_type"):
                row["rfp_type"] = r["rfp_type"]
            clean_batch.append(row)

        supabase.table("taxonomy_v2").insert(clean_batch).execute()
        inserted += len(clean_batch)
        print(f"  {inserted}/{len(records)} 삽입 완료")

    print(f"\ntaxonomy_v2 시드 완료: {inserted}개 행 삽입")

    # 검증
    result = supabase.table("taxonomy_v2").select("level", count="exact").execute()
    print(f"DB 총 행 수: {result.count}")

    # PR 템플릿 그룹 분포
    l3_records = [r for r in records if r["level"] == 3]
    pr_groups = {}
    for r in l3_records:
        key = r.get("pr_template_key", "_generic")
        pr_groups[key] = pr_groups.get(key, 0) + 1
    print(f"\nPR 템플릿 그룹 분포 ({len(pr_groups)}개 그룹):")
    for k, v in sorted(pr_groups.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}개 L3")


if __name__ == "__main__":
    run()
