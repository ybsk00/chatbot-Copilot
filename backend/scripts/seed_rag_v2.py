"""0318 Excel 기반 구조화된 RAG 청크 시드 — LLM 최적화 마크다운 형식

Excel(차세대_품목체계_설명_0318.xlsx)에서 151개 L3 데이터를 읽어
LLM이 섹션별로 정보를 추출하기 좋은 마크다운 청크로 변환하여
knowledge_chunks 테이블에 임베딩과 함께 삽입한다.

사용법:
  cd backend
  python -m scripts.seed_rag_v2
  python -m scripts.seed_rag_v2 --dry-run   (DB 저장 없이 청크 미리보기)
"""
import sys
import os
import io
import time
import pathlib
import argparse

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from dotenv import load_dotenv
load_dotenv()

import openpyxl
from app.db.supabase_client import get_client
from app.rag.embedder import embed_document


# ═══════════════════════════════════════════
# Excel 파일 탐색
# ═══════════════════════════════════════════
def find_excel() -> pathlib.Path | None:
    """Windows 한글 경로 호환 — pathlib으로 0318 Excel 파일 탐색"""
    for depth in range(2, 5):
        base = pathlib.Path(__file__).resolve()
        for _ in range(depth):
            base = base.parent
        for f in base.glob("*.xlsx"):
            if "0318" in f.name:
                return f
    return None


# ═══════════════════════════════════════════
# Excel 파싱 — L3 레코드 + L1/L2 이름 추출
# ═══════════════════════════════════════════
def parse_excel(filepath: pathlib.Path) -> list[dict]:
    """Excel 파싱 → L3 레코드 리스트 (L1/L2 이름 포함)"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[0]

    records = []
    current_l1_name = ""
    current_l2_name = ""

    # 0318은 7열: L1대분류 | L2중분류 | L3소분류 | 공급사 | 전략 | 설명 | 키워드
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        cols = list(row[:7]) if len(row) >= 7 else list(row) + [None] * (7 - len(row))
        l1_name, l2_name, l3_name, suppliers, strategy, description, keywords = cols

        # L1 이름 갱신
        if l1_name:
            current_l1_name = str(l1_name).strip()

        # L2 이름 갱신
        if l2_name:
            current_l2_name = str(l2_name).strip()

        # L3 레코드만 수집
        if not l3_name:
            continue

        l3_name_str = str(l3_name).strip()
        if not l3_name_str:
            continue

        # 공급사 파싱
        sup_str = ""
        if suppliers:
            sup_list = [s.strip() for s in str(suppliers).split(",") if s.strip()]
            sup_str = ", ".join(sup_list)

        # 키워드 파싱
        kw_str = ""
        if keywords:
            kw_list = [k.strip() for k in str(keywords).split(",") if k.strip()]
            kw_str = ", ".join(kw_list)

        # 구매 전략
        strategy_str = str(strategy).strip() if strategy else ""

        # 품목 설명
        desc_str = str(description).strip() if description else ""

        records.append({
            "l1": current_l1_name,
            "l2": current_l2_name,
            "l3": l3_name_str,
            "description": desc_str,
            "purchase_strategy": strategy_str,
            "suppliers": sup_str,
            "keywords": kw_str,
        })

    return records


# ═══════════════════════════════════════════
# 구조화된 청크 텍스트 생성 (LLM 최적화 마크다운)
# ═══════════════════════════════════════════
def build_chunk_text(record: dict) -> str:
    """L3 레코드를 LLM이 잘 인식하는 마크다운 청크로 변환"""
    parts = [
        f"# {record['l3']}",
        f"분류: {record['l1']} > {record['l2']} > {record['l3']}",
    ]

    # 품목 설명
    parts.append("")
    parts.append("## 품목 설명")
    parts.append(record["description"] if record["description"] else "설명 없음")

    # 구매 전략
    parts.append("")
    parts.append("## 구매 전략")
    parts.append(record["purchase_strategy"] if record["purchase_strategy"] else "전략 정보 없음")

    # 주요 공급사
    parts.append("")
    parts.append("## 주요 공급사")
    parts.append(record["suppliers"] if record["suppliers"] else "공급사 정보 없음")

    # 관련 키워드
    parts.append("")
    parts.append("## 관련 키워드")
    parts.append(record["keywords"] if record["keywords"] else "키워드 없음")

    return "\n".join(parts)


# ═══════════════════════════════════════════
# 기존 청크 삭제 (category != 'faq')
# ═══════════════════════════════════════════
def delete_existing_chunks(supabase) -> int:
    """knowledge_chunks에서 category != 'faq'인 일반 청크 전부 삭제"""
    # 기존 non-faq 청크 조회
    existing = supabase.table("knowledge_chunks").select("id").neq("category", "faq").execute()
    if not existing.data:
        return 0

    ids = [r["id"] for r in existing.data]
    # 50개씩 배치 삭제
    for i in range(0, len(ids), 50):
        batch_ids = ids[i:i + 50]
        supabase.table("knowledge_chunks").delete().in_("id", batch_ids).execute()

    return len(ids)


# ═══════════════════════════════════════════
# 메인 실행
# ═══════════════════════════════════════════
def run():
    # Windows cp949 인코딩 문제 방지
    if sys.platform == "win32":
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace", line_buffering=True)

    parser = argparse.ArgumentParser(description="0318 Excel → 구조화된 RAG 청크 시드")
    parser.add_argument("--dry-run", action="store_true", help="DB 저장 없이 청크 미리보기")
    args = parser.parse_args()

    # 1. Excel 파일 탐색
    excel_path = find_excel()
    if not excel_path or not excel_path.exists():
        print("[X] Excel 파일을 찾을 수 없습니다 (차세대_품목체계_설명_0318.xlsx)")
        sys.exit(1)
    print(f"[*] Excel 파일: {excel_path.name}")

    # 2. Excel 파싱
    records = parse_excel(excel_path)
    print(f"[*] 파싱 완료: L3 {len(records)}개 레코드")

    if args.dry_run:
        print("\n[DRY RUN] 처음 3개 청크 미리보기:\n")
        for i, rec in enumerate(records[:3]):
            text = build_chunk_text(rec)
            print(f"--- 청크 {i+1} ---")
            print(f"category: {rec['l1']}")
            print(f"doc_name: {rec['l1']}/{rec['l2']}/{rec['l3']}")
            print(text)
            print()
        print(f"[DRY RUN] 총 {len(records)}개 청크 생성 예정 (DB 저장 안 함)")
        return

    # 3. DB 연결 + 기존 청크 삭제
    supabase = get_client()

    deleted = delete_existing_chunks(supabase)
    print(f"[*] 기존 청크 {deleted}개 삭제 완료 (category != 'faq')")

    # 4. 임베딩 + 삽입
    inserted = 0
    errors = 0

    for idx, rec in enumerate(records):
        chunk_text = build_chunk_text(rec)
        doc_name = f"{rec['l1']}/{rec['l2']}/{rec['l3']}"

        try:
            embedding = embed_document(chunk_text)

            supabase.table("knowledge_chunks").insert({
                "content": chunk_text,
                "embedding": embedding,
                "category": rec["l1"],
                "doc_name": doc_name,
                "chunk_index": 0,
                "metadata": {
                    "l1": rec["l1"],
                    "l2": rec["l2"],
                    "l3": rec["l3"],
                    "source": "taxonomy_v2",
                    "version": "0318",
                },
            }).execute()
            inserted += 1

            if (idx + 1) % 10 == 0:
                print(f"  {idx + 1}/{len(records)} 완료 ({inserted} 삽입, {errors} 오류)")

            # Rate limit 보호 (gemini embedding API)
            time.sleep(0.15)

        except Exception as e:
            errors += 1
            print(f"  [오류] {rec['l3']}: {e}")
            time.sleep(1)

    print(f"\n[DONE] RAG 시드 완료: {inserted}개 삽입, {errors}개 오류")
    print(f"  knowledge_chunks에 L1 대분류명을 category로 저장됨")
    print(f"  metadata.source='taxonomy_v2', metadata.version='0318'")


if __name__ == "__main__":
    run()
