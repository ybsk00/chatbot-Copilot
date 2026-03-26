"""PR(구매요청서) 템플릿 v3 시드 — 표준 구매요청서 139개 L3 개별 템플릿
Excel 파일에서 자동 파싱하여 pr_templates 테이블에 시드
"""
import sys, os, re, json, glob
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from dotenv import load_dotenv
load_dotenv()

import openpyxl
from app.db.supabase_client import get_client

# ── 프로젝트 루트 (표준 구매요청서 폴더 위치) ──
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
PR_EXCEL_DIR = os.path.join(PROJECT_ROOT, "표준 구매요청서")

# ── 공통 필드 (c1~c20, 기존과 동일) ──
COMMON_LABELS = {
    "c1": "발주기관명", "c2": "요청부서", "c3": "요청자", "c4": "연락처", "c5": "이메일",
    "c6": "서비스/품목명", "c7": "구매/계약 목적", "c8": "계약 유형",
    "c9": "계약 기간", "c10": "대상 규모/수량",
    "c11": "서비스 범위/요구 사양", "c12": "제공/수행 방식",
    "c13": "품질/SLA 기준", "c14": "보안/법적 요건",
    "c15": "단가 산정 방식", "c16": "구간 할인 적용", "c17": "결제 주기 및 방식",
    "c18": "단가 인상/에스컬레이션 조건", "c19": "계약 해지/위약금 조건",
    "c20": "BSM/전산 연동 조건",
}
COMMON_REQUIRED = {
    "c1": True, "c2": True, "c3": True, "c4": True, "c5": True,
    "c6": True, "c7": True, "c8": True, "c9": True, "c10": True,
    "c11": True, "c12": True, "c13": False, "c14": False,
    "c15": False, "c16": False, "c17": False,
    "c18": False, "c19": False, "c20": False,
}
COMMON_SECTIONS_PREFIX = [
    {"title": "1. 요청자 정보", "fields": ["c1", "c2", "c3", "c4", "c5"], "icon": "org"},
    {"title": "2. 계약 기본 정보", "fields": ["c6", "c7", "c8", "c9", "c10"], "icon": "doc"},
    {"title": "3. 서비스 요건", "fields": ["c11", "c12", "c13", "c14"], "icon": "gear"},
]
PAYMENT_SECTION_FIELDS = ["c15", "c16", "c17", "c18", "c19", "c20"]

# ── L1 폴더 → 대분류명 매핑 ──
L1_FOLDER_MAP = {
    "L01": "사무·총무", "L02": "인사·복리후생", "L03": "시설·건물관리",
    "L04": "차량·출장", "L05": "보험 서비스", "L06": "전문용역·컨설팅",
    "L07": "마케팅", "L08": "IT/ICT", "L09": "물류",
    "L10": "생산관리", "L11": "연구개발",
}

# Row 6의 컨텍스트별 필드 → c 필드 매핑 힌트
ROW6_HINTS = {
    "렌탈": {"c9": "렌탈 기간", "c10": "설치 대수/사업장 수"},
    "계약": {"c9": "계약 기간", "c10": "대상 규모"},
    "도입": {"c9": "도입 시기", "c10": "수량/규모"},
    "캠페인": {"c9": "캠페인 기간", "c10": "타겟/규모"},
    "시작일": {"c9": "시작일", "c10": "수량/규모"},
}


def build_common_fields(defaults: dict = None) -> dict:
    defaults = defaults or {}
    fields = {}
    for key, label in COMMON_LABELS.items():
        d = defaults.get(key, "")
        req = COMMON_REQUIRED[key]
        f = {"label": label, "value": "", "required": req}
        if d:
            f["default"] = d
        fields[key] = f
    return fields


def extract_prefix(filename: str) -> str | None:
    """파일명에서 L3 prefix 추출: L1101, M1001, R2001 등"""
    m = re.match(r'^([LMR]\d{4})', filename)
    return m.group(1) if m else None


def extract_name(filename: str, prefix: str) -> str:
    """파일명에서 서비스명 추출"""
    name = filename.replace(prefix + "_", "").replace("_구매요청서.xlsx", "")
    name = name.replace("_", " ").strip()
    return name


def parse_pr_excel(filepath: str) -> dict | None:
    """표준 구매요청서 Excel 파싱 → PR 템플릿 dict"""
    filename = os.path.basename(filepath)
    prefix = extract_prefix(filename)
    if not prefix:
        print(f"  [SKIP] prefix 추출 실패: {filename}")
        return None

    name = extract_name(filename, prefix)

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  [ERROR] {filename}: {e}")
        return None

    ws = wb[wb.sheetnames[0]]  # Sheet 1: 구매요청서

    # ── Row 5-6: 헤더 필드에서 c 기본값 추출 ──
    c_defaults = {}
    # Row 2: 제목에서 서비스명 추출
    title_val = ws.cell(2, 1).value
    if title_val:
        # 이모지 제거 후 서비스명
        clean_title = re.sub(r'[📎📋🔧📊💰💹📖]', '', str(title_val)).strip()
        if clean_title:
            # "정수기·공기청정기·비데 렌탈 서비스 구매요청서" → "정수기·공기청정기·비데 렌탈 서비스"
            c_defaults["c6"] = clean_title.split("구매요청서")[0].strip()

    # Row 6 C4: 컨텍스트별 날짜/기간 필드
    row6_c4 = ws.cell(6, 4).value
    if row6_c4:
        r6str = str(row6_c4)
        for hint_key, hint_map in ROW6_HINTS.items():
            if hint_key in r6str:
                c_defaults.update(hint_map)
                break

    # ── Row 9+: 항목 테이블 파싱 ──
    p_fields = {}       # p1, p2, ... → field dict
    sections = []       # [{title, fields, icon}]
    current_section = None
    current_section_fields = []
    p_idx = 1

    for row in range(9, ws.max_row + 1):
        c1 = ws.cell(row, 1).value  # No. or ▶ header
        c2 = ws.cell(row, 2).value  # 분류
        c3 = ws.cell(row, 3).value  # 항목명
        c4 = ws.cell(row, 4).value  # 항목 설명
        c5 = ws.cell(row, 5).value  # 필수/옵션

        if c1 is None and c2 is None and c3 is None:
            continue

        c1_str = str(c1 or "").strip()

        # ▶ 섹션 헤더 감지
        if c1_str.startswith("▶"):
            # 이전 섹션 저장
            if current_section and current_section_fields:
                sections.append({
                    "title": current_section,
                    "fields": current_section_fields[:],
                    "icon": "gear",
                })
            # 새 섹션 시작
            section_name = c1_str.replace("▶", "").strip()
            # [공통] 도입 기본 정보 → 공통 - 도입 기본 정보
            section_name = re.sub(r'\[(.+?)\]\s*', r'\1 - ', section_name)
            current_section = section_name
            current_section_fields = []
            continue

        # 데이터 행 (No.가 숫자)
        try:
            item_no = int(c1_str)
        except (ValueError, TypeError):
            continue

        if not c3:
            continue

        # 필수/옵션 판정
        c5_str = str(c5 or "").strip().lower()
        required = "필수" in c5_str

        # p 필드 생성
        field_key = f"p{p_idx}"
        field = {
            "label": str(c3).strip(),
            "value": "",
            "required": required,
        }

        # 항목 설명 → description (너무 길면 자름)
        if c4:
            desc = str(c4).replace("\n", " ").strip()[:200]
            field["description"] = desc

        # 요청자 입력값 (col H=8) → default
        c8 = ws.cell(row, 8).value
        if c8:
            default_val = str(c8).replace("\n", " ").strip()[:100]
            field["default"] = default_val

        p_fields[field_key] = field
        current_section_fields.append(field_key)
        p_idx += 1

    # 마지막 섹션 저장
    if current_section and current_section_fields:
        sections.append({
            "title": current_section,
            "fields": current_section_fields[:],
            "icon": "gear",
        })

    wb.close()

    if not p_fields:
        print(f"  [WARN] 항목 없음: {filename}")
        return None

    # ── 섹션 넘버링 (공통 3개 + 고유 N개 + 결제 1개) ──
    numbered_sections = COMMON_SECTIONS_PREFIX[:]
    for i, s in enumerate(sections):
        numbered_sections.append({
            "title": f"{i + 4}. {s['title']}",
            "fields": s["fields"],
            "icon": s.get("icon", "gear"),
        })
    payment_idx = len(sections) + 4
    numbered_sections.append({
        "title": f"{payment_idx}. 결제·계약 조건",
        "fields": PAYMENT_SECTION_FIELDS,
        "icon": "mail",
    })

    # ── 최종 템플릿 dict ──
    all_fields = {**build_common_fields(c_defaults), **p_fields}

    return {
        "type_key": prefix,
        "name": name,
        "description": f"{name} 구매요청서 ({len(p_fields)}개 항목)",
        "fields": all_fields,
        "sections": numbered_sections,
    }


def find_l1_folder(folder_name: str) -> str | None:
    """폴더명에서 L1 코드 추출"""
    m = re.match(r'^(L\d{2})', folder_name)
    return m.group(1) if m else None


def run():
    supabase = get_client()

    # ── Excel 파일 수집 ──
    templates = []
    errors = []

    if not os.path.isdir(PR_EXCEL_DIR):
        print(f"[ERROR] 폴더 없음: {PR_EXCEL_DIR}")
        return

    l1_folders = sorted([d for d in os.listdir(PR_EXCEL_DIR)
                         if os.path.isdir(os.path.join(PR_EXCEL_DIR, d))])

    print(f"=== 표준 구매요청서 폴더 {len(l1_folders)}개 ===")

    for folder in l1_folders:
        folder_path = os.path.join(PR_EXCEL_DIR, folder)
        l1_code = find_l1_folder(folder)
        l1_name = L1_FOLDER_MAP.get(l1_code, folder) if l1_code else folder

        xlsx_files = sorted(glob.glob(os.path.join(folder_path, "*.xlsx")))
        print(f"\n[{folder}] ({l1_name}) — {len(xlsx_files)}개 파일")

        for fpath in xlsx_files:
            tpl = parse_pr_excel(fpath)
            if tpl:
                tpl["category_group"] = l1_name
                templates.append(tpl)
                req_count = sum(1 for f in tpl["fields"].values() if f.get("required"))
                opt_count = len(tpl["fields"]) - req_count
                p_count = sum(1 for k in tpl["fields"] if k.startswith("p"))
                print(f"  [OK] {tpl['type_key']}: {tpl['name']} "
                      f"(p필드 {p_count}개, 필수 {req_count}, 옵션 {opt_count})")
            else:
                errors.append(os.path.basename(fpath))

    print(f"\n=== 파싱 완료: {len(templates)}개 성공, {len(errors)}개 실패 ===")
    if errors:
        print(f"  실패: {errors}")

    if not templates:
        print("[ERROR] 파싱된 템플릿 없음. 시드 중단.")
        return

    # ── _generic 폴백 추가 ──
    generic_tpl = {
        "type_key": "_generic",
        "name": "일반 구매요청",
        "description": "범용 구매요청서 (분류 미매칭 시 사용)",
        "category_group": "기타",
        "fields": {
            **build_common_fields(),
            "p1": {"label": "요청 품목/서비스명", "value": "", "required": True},
            "p2": {"label": "상세 요구사항", "value": "", "required": True},
            "p3": {"label": "납품/서비스 시작 희망일", "value": "", "required": True},
            "p4": {"label": "예상 예산", "value": "", "required": False},
            "p5": {"label": "기타 요청사항", "value": "", "required": False},
        },
        "sections": COMMON_SECTIONS_PREFIX + [
            {"title": "4. 요청 내용", "fields": ["p1", "p2", "p3", "p4", "p5"], "icon": "gear"},
            {"title": "5. 결제·계약 조건", "fields": PAYMENT_SECTION_FIELDS, "icon": "mail"},
        ],
    }
    templates.append(generic_tpl)

    # ── DB 시드 ──
    print(f"\n=== DB 시드 시작 ({len(templates)}개) ===")

    # 기존 삭제
    supabase.table("pr_templates").delete().neq("id", 0).execute()
    print("기존 PR 템플릿 삭제 완료")

    success = 0
    for t in templates:
        try:
            supabase.table("pr_templates").insert({
                "type_key": t["type_key"],
                "name": t["name"],
                "description": t["description"],
                "fields": t["fields"],
                "sections": t["sections"],
                "is_active": True,
            }).execute()
            success += 1
        except Exception as e:
            print(f"  [DB ERROR] {t['type_key']}: {e}")

    print(f"\n=== PR 템플릿 v3 시드 완료: {success}/{len(templates)}개 ===")

    # ── 통계 ──
    print("\n=== 템플릿별 필드 통계 ===")
    for t in templates[:5]:  # 처음 5개만 표시
        total = len(t["fields"])
        req = sum(1 for f in t["fields"].values() if f.get("required"))
        p_cnt = sum(1 for k in t["fields"] if k.startswith("p"))
        print(f"  {t['type_key']}: 총 {total}개 (공통 20 + 고유 {p_cnt}) — 필수 {req}")
    if len(templates) > 5:
        print(f"  ... 외 {len(templates) - 5}개")


if __name__ == "__main__":
    run()
