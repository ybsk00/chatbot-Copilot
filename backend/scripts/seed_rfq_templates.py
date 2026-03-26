"""RFQ(견적서) 템플릿 시드 — 표준 견적서 37개 L3 개별 템플릿
Excel 파일에서 소싱담당자 영역(Zone 1-2)만 추출하여 rfq_templates 테이블에 시드
시장렌탈참고가, B2B예산단가 컬럼 제외
"""
import sys, os, re, glob
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from dotenv import load_dotenv
load_dotenv()

import openpyxl
from app.db.supabase_client import get_client

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
RFQ_EXCEL_DIR = os.path.join(PROJECT_ROOT, "표준 견적서")

# ── 제외 컬럼 키워드 ──
EXCLUDE_KEYWORDS = ["시장렌탈료참고", "B2B예산단가참고", "시장렌탈참고", "B2B예산", "시장 참고"]

# ── L1 폴더 매핑 ──
L1_FOLDER_MAP = {
    "L01": "사무·총무", "L02": "인사·복리후생", "L03": "시설·건물관리",
}

# ── 시트별 섹션 매핑 ──
SHEET_SECTION_MAP = {
    0: {"prefix": "가격", "icon": "money"},    # 품목 가격 견적
    1: {"prefix": "계약", "icon": "doc"},       # 계약 조건
    2: {"prefix": "부가서비스", "icon": "gear"}, # 부가 서비스
    3: {"prefix": "SLA", "icon": "chart"},      # SLA·품질 관리
    4: {"prefix": "TCO", "icon": "calc"},       # TCO·정산 합계
}


def extract_prefix(filename: str) -> str | None:
    m = re.match(r'^([LMR]\d{4})', filename)
    return m.group(1) if m else None


def extract_name(filename: str, prefix: str) -> str:
    name = filename.replace(prefix + "_", "")
    name = re.sub(r'_표준견적서.*\.xlsx$', '', name)
    name = name.replace("_", " ").strip()
    return name


def is_excluded_col(header: str) -> bool:
    """제외 컬럼인지 확인"""
    if not header:
        return False
    h = str(header).strip()
    return any(kw in h for kw in EXCLUDE_KEYWORDS)


def is_buyer_zone(header_row: list, col_idx: int, zone_markers: dict) -> bool:
    """Zone 1-2 (소싱담당자 영역) 여부 판단"""
    # zone_markers: {col_idx: zone_type}
    # Zone 1 (🔵): 소싱담당자 필수
    # Zone 2 (🔷): 소싱담당자 선택
    # Zone 3 (🟠): 공급사 필수 → 제외
    # Zone 4 (🟢): 공급사 선택 → 제외

    for zm_col, zm_type in sorted(zone_markers.items(), reverse=True):
        if col_idx >= zm_col:
            return zm_type in ("buyer_required", "buyer_optional")
    return True  # 기본: 포함


def detect_zone_markers(ws, marker_row: int) -> dict:
    """Row 8 등에서 zone 마커 위치 감지"""
    markers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(marker_row, c).value
        if not v:
            continue
        vs = str(v).strip()
        if "소싱담당자" in vs and "필수" in vs:
            markers[c] = "buyer_required"
        elif "소싱담당자" in vs and "선택" in vs:
            markers[c] = "buyer_optional"
        elif "공급사" in vs and "필수" in vs:
            markers[c] = "supplier_required"
        elif "공급사" in vs and "선택" in vs:
            markers[c] = "supplier_optional"
    return markers


def parse_sheet(ws, sheet_idx: int) -> list:
    """시트에서 소싱담당자 Zone 1-2 항목만 추출"""
    items = []

    # Zone 마커 감지 (Row 3 or Row 8)
    zone_markers = {}
    for try_row in [3, 8]:
        zm = detect_zone_markers(ws, try_row)
        if zm:
            zone_markers = zm
            break

    # 헤더 행 찾기 (Row 9가 표준)
    header_row = 9
    for r in range(7, min(ws.max_row + 1, 12)):
        c1 = ws.cell(r, 1).value
        if c1 and str(c1).strip() == "No.":
            header_row = r
            break

    # 헤더 읽기
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v:
            headers[c] = str(v).strip()

    # 데이터 행 파싱
    for r in range(header_row + 1, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        if c1 is None:
            continue

        c1_str = str(c1).strip()

        # ▶ 섹션 헤더나 합계행, 서명행 스킵
        if c1_str.startswith("▶") or c1_str.startswith("🔶") or c1_str.startswith("발주사"):
            continue

        # No.가 숫자 또는 ①②③ 등
        try:
            int(c1_str)
        except ValueError:
            if not re.match(r'^[①②③④⑤⑥⑦⑧⑨⑩]', c1_str):
                continue

        # 소싱담당자 영역 컬럼만 추출
        item_data = {}
        for c, header in headers.items():
            if c == 1:  # No. 컬럼 스킵
                continue
            if is_excluded_col(header):
                continue
            if zone_markers and not is_buyer_zone(None, c, zone_markers):
                continue

            val = ws.cell(r, c).value
            if val is not None:
                item_data[header] = str(val).strip()[:100]

        if not item_data:
            continue

        # 항목명 결정 (각 시트마다 다름)
        label = ""
        for key in ["항  목  명", "항목", "제품 유형", "비용 항목", "SLA 항목", "서비스 구분"]:
            if key in item_data:
                label = item_data[key]
                break
        if not label:
            # 첫 번째 비어있지 않은 값 사용
            label = next(iter(item_data.values()), f"항목 {c1_str}")

        # 필수/선택 판정
        required = True
        for key in ["필수/선택", "필수/ 선택"]:
            if key in item_data:
                required = "필수" in item_data[key]
                break

        # 설명 구성
        desc_parts = []
        for key in ["발주사 요건 ★필수", "세부 설명", "제품 사양", "산출 기준",
                     "최소 요건 ★필수", "요구 기능 ★필수", "항목 상세 설명"]:
            if key in item_data:
                desc_parts.append(item_data[key])
        description = " | ".join(desc_parts)[:200] if desc_parts else ""

        items.append({
            "label": label,
            "required": required,
            "description": description,
            "raw": item_data,
        })

    return items


def parse_rfq_excel(filepath: str) -> dict | None:
    """표준 견적서 Excel 파싱 → RFQ 템플릿 dict"""
    filename = os.path.basename(filepath)
    prefix = extract_prefix(filename)
    if not prefix:
        return None

    name = extract_name(filename, prefix)

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  [ERROR] {filename}: {e}")
        return None

    q_fields = {}
    sections = []
    q_idx = 1

    # Sheet 1~5 파싱 (Sheet 6 작성가이드 skip)
    for si in range(min(5, len(wb.sheetnames))):
        sheet_name = wb.sheetnames[si]
        ws = wb[sheet_name]
        section_info = SHEET_SECTION_MAP.get(si, {"prefix": f"시트{si+1}", "icon": "gear"})

        items = parse_sheet(ws, si)
        if not items:
            continue

        section_fields = []
        for item in items:
            fkey = f"q{q_idx}"
            q_fields[fkey] = {
                "label": item["label"],
                "value": "",
                "required": item["required"],
            }
            if item["description"]:
                q_fields[fkey]["description"] = item["description"]

            section_fields.append(fkey)
            q_idx += 1

        if section_fields:
            sections.append({
                "title": f"{len(sections) + 1}. {section_info['prefix']}",
                "fields": section_fields,
                "icon": section_info["icon"],
            })

    wb.close()

    if not q_fields:
        print(f"  [WARN] 항목 없음: {filename}")
        return None

    return {
        "type_key": prefix,
        "name": name,
        "description": f"{name} 견적서 ({len(q_fields)}개 항목)",
        "fields": q_fields,
        "sections": sections,
    }


def run():
    supabase = get_client()

    if not os.path.isdir(RFQ_EXCEL_DIR):
        print(f"[ERROR] 폴더 없음: {RFQ_EXCEL_DIR}")
        return

    templates = []
    l1_folders = sorted([d for d in os.listdir(RFQ_EXCEL_DIR)
                         if os.path.isdir(os.path.join(RFQ_EXCEL_DIR, d))])

    print(f"=== 표준 견적서 폴더 {len(l1_folders)}개 ===")

    for folder in l1_folders:
        folder_path = os.path.join(RFQ_EXCEL_DIR, folder)
        l1_code = re.match(r'^(L\d{2})', folder)
        l1_name = L1_FOLDER_MAP.get(l1_code.group(1), folder) if l1_code else folder

        xlsx_files = sorted(glob.glob(os.path.join(folder_path, "*.xlsx")))
        print(f"\n[{folder}] ({l1_name}) — {len(xlsx_files)}개 파일")

        for fpath in xlsx_files:
            tpl = parse_rfq_excel(fpath)
            if tpl:
                tpl["category_group"] = l1_name
                templates.append(tpl)
                q_count = len(tpl["fields"])
                req = sum(1 for f in tpl["fields"].values() if f.get("required"))
                print(f"  [OK] {tpl['type_key']}: {tpl['name']} "
                      f"(q필드 {q_count}개, 필수 {req}, 옵션 {q_count - req}, 섹션 {len(tpl['sections'])})")
            else:
                print(f"  [SKIP] {os.path.basename(fpath)}")

    # 중복 제거 (L1301 typo 등)
    seen = set()
    deduped = []
    for t in templates:
        if t["type_key"] not in seen:
            seen.add(t["type_key"])
            deduped.append(t)
    templates = deduped

    print(f"\n=== 파싱 완료: {len(templates)}개 ===")

    if not templates:
        print("[ERROR] 파싱된 템플릿 없음. 시드 중단.")
        return

    # DB 시드
    print(f"\n=== DB 시드 시작 ({len(templates)}개) ===")
    supabase.table("rfq_templates").delete().neq("id", 0).execute()
    print("기존 RFQ 템플릿 삭제 완료")

    success = 0
    for t in templates:
        try:
            supabase.table("rfq_templates").insert({
                "type_key": t["type_key"],
                "name": t["name"],
                "description": t["description"],
                "category_group": t.get("category_group", ""),
                "fields": t["fields"],
                "sections": t["sections"],
                "is_active": True,
            }).execute()
            success += 1
        except Exception as e:
            print(f"  [DB ERROR] {t['type_key']}: {e}")

    print(f"\n=== RFQ 템플릿 시드 완료: {success}/{len(templates)}개 ===")


if __name__ == "__main__":
    run()
